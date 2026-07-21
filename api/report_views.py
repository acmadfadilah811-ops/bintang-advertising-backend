"""Modul Laporan — fondasi generik + handler per laporan.

Pola: setiap laporan mendaftar lewat dekorator @report(...) dengan id yang sama
dengan `id` pada definisi frontend (src/features/laporan/pages/*/reportList*.js),
dan mengembalikan {'rows': [...], 'summary': {...}} dengan key baris PERSIS sama
dengan `columns[].key` di definisi tersebut.

Menambah laporan baru = tambah satu handler di sini + satu baris `dataSource`
di definisi frontend. Endpoint data dan export otomatis ikut bekerja.
"""
import io

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from datetime import timedelta
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework.permissions import IsAuthenticated
from .throttles import ReportRateThrottle
from rest_framework.response import Response
from rest_framework.views import APIView

from .export_views import IgnoreFormatContentNegotiation
from .models import Order
from .pos_models import POSSale
from .product_models import (
    Product, ProductStockMovement,
    StockInDocument, StockInDocumentItem,
    StockOutDocument, StockOutDocumentItem,
    StockOpnameDocument, StockOpnameDocumentItem,
    StockLayerConsumption,
)

# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------
REPORT_REGISTRY = {}


def report(report_id, label, columns):
    """Daftarkan handler laporan.

    columns: list of (key, label, type) — type ∈ 'text'|'qty'|'money'|'date'.
    Dipakai untuk header export DAN dikirim ke frontend sebagai petunjuk format,
    sehingga `rows` bisa berisi angka mentah (bagus untuk Excel) tapi tetap
    tampil terformat di layar.
    """
    def deco(fn):
        REPORT_REGISTRY[report_id] = {
            'handler': fn,
            'label': label,
            'columns': [{'key': k, 'label': l, 'type': t} for k, l, t in columns],
        }
        return fn
    return deco


def _params(request):
    q = request.query_params
    return {
        'start': parse_date(q.get('start')) if q.get('start') else None,
        'end': parse_date(q.get('end')) if q.get('end') else None,
        'search': (q.get('search') or '').strip(),
        'sort': q.get('sort') or '',
        'extra': q,
    }


def _num(v):
    return float(v or 0)


def _sku_rows():
    """Universe produk+varian sebagai baris datar (dipakai laporan stok)."""
    out = []
    products = Product.objects.all().select_related('kategori', 'brand').prefetch_related('variants')
    for p in products:
        variants = list(p.variants.all()) if p.has_variant else []
        if variants:
            for v in variants:
                out.append({
                    'product': p, 'variant': v,
                    'nama': p.nama, 'varian': v.nama_varian,
                    'sku': v.sku or p.sku or '',
                    'kategori': p.kategori.nama if p.kategori else 'Umum',
                    'brand': p.brand.nama if p.brand else '',
                    'qty': _num(v.qty_stok),
                    'harga_beli': _num(v.harga_beli or p.harga_beli),
                    'jual_online': _num(v.harga_jual_online or p.harga_jual_online),
                    'jual_toko': _num(v.harga_jual_toko or p.harga_jual_toko),
                })
        else:
            out.append({
                'product': p, 'variant': None,
                'nama': p.nama, 'varian': '',
                'sku': p.sku or '',
                'kategori': p.kategori.nama if p.kategori else 'Umum',
                'brand': p.brand.nama if p.brand else '',
                'qty': _num(p.qty_stok),
                'harga_beli': _num(p.harga_beli),
                'jual_online': _num(p.harga_jual_online),
                'jual_toko': _num(p.harga_jual_toko),
            })
    return out


def _last_supplier_map():
    """product_id -> nama supplier dari dokumen Stok Masuk terakhir."""
    m = {}
    qs = (
        StockInDocumentItem.objects
        .select_related('document')
        .order_by('document__tanggal', 'id')
    )
    for it in qs:
        if it.document and it.document.supplier:
            m[it.product_id] = it.document.supplier
    return m


def _date_filter_docs(qs, params, field='tanggal'):
    if params['start']:
        qs = qs.filter(**{f'{field}__gte': params['start']})
    if params['end']:
        qs = qs.filter(**{f'{field}__lte': params['end']})
    return qs


# ---------------------------------------------------------------------------
# Laporan Produk — stok
# ---------------------------------------------------------------------------
@report('sisa-produk', 'Sisa Stok Produk', [
    ('produk', 'Produk', 'text'), ('variant', 'Variant', 'text'),
    ('kategori', 'Kategori', 'text'), ('qty_stok', 'Qty Stok', 'qty'),
    ('harga_beli', 'Harga Beli', 'money'), ('subtotal', 'Subtotal', 'money'),
    ('harga_jual_online', 'Harga Jual Online', 'money'),
    ('harga_jual_toko', 'Harga Jual di Toko', 'money'),
])
def rpt_sisa_produk(params):
    rows = []
    tot_qty = tot_beli = tot_online = tot_toko = 0.0
    for s in _sku_rows():
        if params['search'] and params['search'].lower() not in f"{s['nama']} {s['sku']}".lower():
            continue
        subtotal = s['qty'] * s['harga_beli']
        rows.append({
            'produk': s['nama'], 'variant': s['varian'], 'kategori': s['kategori'],
            'qty_stok': s['qty'], 'harga_beli': s['harga_beli'], 'subtotal': subtotal,
            'harga_jual_online': s['jual_online'], 'harga_jual_toko': s['jual_toko'],
        })
        tot_qty += s['qty']
        tot_beli += subtotal
        tot_online += s['qty'] * s['jual_online']
        tot_toko += s['qty'] * s['jual_toko']
    return {
        'rows': rows,
        'summary': {'rows': [{
            'qty_stok': tot_qty, 'total_harga_beli': tot_beli,
            'total_jual_online': tot_online, 'total_jual_toko': tot_toko,
        }]},
    }


@report('produk-status', 'Produk Status', [
    ('produk', 'Produk', 'text'), ('kategori', 'Kategori', 'text'),
    ('qty_stok', 'Qty Stok', 'qty'), ('harga_beli', 'Harga Beli', 'money'),
    ('harga_jual_online', 'Harga Jual Online', 'money'),
    ('harga_jual_toko', 'Harga Jual di Toko', 'money'),
    ('status', 'Status', 'text'),
])
def rpt_produk_status(params):
    rows = []
    for s in _sku_rows():
        p = s['product']
        if not p.is_active:
            status = 'Nonaktif'
        elif s['qty'] <= 0:
            status = 'Habis'
        elif s['qty'] <= _num(p.stok_minimum):
            status = 'Stok Menipis'
        else:
            status = 'Aktif'
        nama = f"{s['nama']} - {s['varian']}" if s['varian'] else s['nama']
        rows.append({
            'produk': nama, 'kategori': s['kategori'], 'qty_stok': s['qty'],
            'harga_beli': s['harga_beli'], 'harga_jual_online': s['jual_online'],
            'harga_jual_toko': s['jual_toko'], 'status': status,
        })
    return {
        'rows': rows,
        'summary': {'rows': [{'jumlah_produk': len(rows)}]},
    }


@report('peringatan-stok', 'Peringatan Sisa Stok', [
    ('produk', 'Produk', 'text'), ('kategori', 'Kategori', 'text'),
    ('supplier', 'Supplier', 'text'), ('qty_stok', 'Qty Stok', 'qty'),
    ('stok_minimum', 'Stok Minimum', 'qty'), ('kekurangan', 'Kekurangan', 'qty'),
    ('harga_beli', 'Harga Beli', 'money'),
])
def rpt_peringatan_stok(params):
    suppliers = _last_supplier_map()
    rows = []
    for s in _sku_rows():
        minimum = _num(s['product'].stok_minimum)
        if minimum <= 0 or s['qty'] > minimum:
            continue
        nama = f"{s['nama']} - {s['varian']}" if s['varian'] else s['nama']
        rows.append({
            'produk': nama, 'kategori': s['kategori'],
            'supplier': suppliers.get(s['product'].id, ''),
            'qty_stok': s['qty'],
            # Dipakai layar "Email Peringatan Stok" dan ringkasan di bawah.
            # Sebelumnya key ini tidak pernah diisi sehingga total_kekurangan selalu 0.
            'stok_minimum': minimum,
            'kekurangan': minimum - s['qty'],
            'harga_beli': s['harga_beli'],
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'produk_perlu_restock': len(rows),
            'total_kekurangan': sum(_num(r.get('kekurangan')) for r in rows),
        }]},
    }


@report('stok-masuk', 'Stok Masuk', [
    ('no_transaksi', 'No. Transaksi', 'text'), ('tanggal', 'Tanggal', 'date'),
    ('produk', 'Produk', 'text'), ('variant', 'Variant', 'text'),
    ('seri', 'Seri', 'text'), ('supplier', 'Supplier', 'text'),
    ('harga_beli', 'Harga Beli', 'money'), ('rata_harga_beli', 'Rata-rata Harga Beli', 'money'),
    ('harga_jual', 'Harga Jual', 'money'), ('subtotal', 'Subtotal', 'money'),
    ('qty', 'Qty', 'qty'),
])
def rpt_stok_masuk(params):
    qs = (StockInDocumentItem.objects
          .select_related('document', 'product', 'variant', 'product__kategori')
          .filter(document__status='selesai'))
    qs = _date_filter_docs(qs, params, 'document__tanggal')
    if params['search']:
        qs = qs.filter(Q(product__nama__icontains=params['search'])
                       | Q(document__nomor__icontains=params['search'])
                       | Q(document__supplier__icontains=params['search']))
    rows = []
    total = 0.0
    for it in qs.order_by('-document__tanggal', '-id'):
        subtotal = _num(it.qty) * _num(it.harga_beli)
        total += subtotal
        rows.append({
            'no_transaksi': it.document.nomor,
            'tanggal': it.document.tanggal.isoformat() if it.document.tanggal else '',
            'produk': it.product.nama if it.product else '',
            'variant': it.variant.nama_varian if it.variant else '',
            'seri': '',
            'supplier': it.document.supplier or '',
            'harga_beli': _num(it.harga_beli),
            'rata_harga_beli': _num(it.product.harga_beli) if it.product else 0,
            'harga_jual': _num(it.product.harga_jual_toko) if it.product else 0,
            'subtotal': subtotal,
            'qty': _num(it.qty),
        })
    return {
        'rows': rows,
        'summary': {'type': 'list', 'items': [
            {'label': 'Total Produk', 'value': len(rows), 'type': 'qty'},
            {'label': 'Jumlah', 'value': total, 'type': 'money'},
        ]},
    }


@report('stok-keluar', 'Stok Keluar', [
    ('no_transaksi', 'No. Transaksi', 'text'), ('tanggal', 'Tanggal', 'date'),
    ('produk', 'Produk', 'text'), ('seri', 'Seri', 'text'),
    ('catatan', 'Catatan', 'text'), ('alasan', 'Alasan', 'text'),
    ('qty', 'Qty', 'qty'), ('harga_beli', 'Harga Beli', 'money'),
    ('subtotal', 'Subtotal', 'money'),
])
def rpt_stok_keluar(params):
    qs = (StockOutDocumentItem.objects
          .select_related('document', 'product', 'variant')
          .filter(document__status='selesai'))
    qs = _date_filter_docs(qs, params, 'document__tanggal')
    if params['search']:
        qs = qs.filter(Q(product__nama__icontains=params['search'])
                       | Q(document__nomor__icontains=params['search']))

    # Filter opsional berdasarkan alasan (dropdown "Alasan" di toolbar).
    alasan_map = {
        'Rusak': 'rusak', 'Kadaluarsa': 'kadaluarsa',
        'Pengembalian dana (Refund)': 'refund',
        'Jumlah stock kelebihan': 'kelebihan_stok',
        'Alasan lainnya': 'lainnya',
    }
    alasan_pilih = params['extra'].get('Alasan') or params['extra'].get('alasan')
    if alasan_pilih and alasan_pilih != 'Semua data':
        kode = alasan_map.get(alasan_pilih, alasan_pilih)
        qs = qs.filter(document__alasan=kode)

    labels = dict(StockOutDocument.REASON_CHOICES)
    rows = []
    total = 0.0
    for it in qs.order_by('-document__tanggal', '-id'):
        harga = _num(it.product.harga_beli) if it.product else 0
        subtotal = _num(it.qty) * harga
        total += subtotal
        rows.append({
            'no_transaksi': it.document.nomor,
            'tanggal': it.document.tanggal.isoformat() if it.document.tanggal else '',
            'produk': it.product.nama if it.product else '',
            'seri': '',
            'catatan': it.document.catatan or '',
            'alasan': labels.get(it.document.alasan, it.document.alasan or ''),
            'qty': _num(it.qty),
            'harga_beli': harga,
            'subtotal': subtotal,
        })
    return {
        'rows': rows,
        'summary': {'type': 'list', 'items': [
            {'label': 'Total Produk', 'value': len(rows), 'type': 'qty'},
            {'label': 'Jumlah', 'value': total, 'type': 'money'},
        ]},
    }


@report('pergerakan-stok', 'Pergerakan Stok', [
    ('tanggal', 'Tanggal', 'date'), ('ref_no', 'Ref No.', 'text'),
    ('produk', 'Produk', 'text'), ('sku', 'SKU', 'text'),
    ('awal', 'Awal', 'qty'), ('masuk', 'Masuk', 'qty'),
    ('keluar', 'Keluar', 'qty'), ('sisa', 'Sisa', 'qty'),
    ('subtotal', 'Subtotal', 'money'),
])
def rpt_pergerakan_stok(params):
    # BE-17: tanpa filter tanggal, endpoint ini memuat SELURUH tabel mutasi ke
    # memori. Terapkan rentang default 30 hari terakhir bila user tidak memberi
    # rentang, sehingga request default tidak mematikan worker.
    start = params.get('start')
    end = params.get('end')
    if not start and not end:
        end = timezone.localdate()
        start = end - timedelta(days=30)

    qs = ProductStockMovement.objects.select_related(
        'product', 'variant', 'stock_in_document', 'stock_out_document',
        'stock_production_document', 'stock_opname_document',
    ).all()
    if start:
        qs = qs.filter(Q(tanggal__gte=start) | Q(tanggal__isnull=True, created_at__date__gte=start))
    if end:
        qs = qs.filter(Q(tanggal__lte=end) | Q(tanggal__isnull=True, created_at__date__lte=end))

    masuk_tipe = {'masuk', 'produksi', 'pengembalian'}
    rows = []
    for m in qs.order_by('-created_at'):
        doc = (m.stock_in_document or m.stock_out_document
               or m.stock_production_document or m.stock_opname_document)
        ref = doc.nomor if doc else ''
        if params['search'] and params['search'].lower() not in ref.lower():
            continue
        qty = _num(m.qty)
        is_masuk = m.tipe in masuk_tipe
        tgl = m.tanggal or (m.created_at.date() if m.created_at else None)
        rows.append({
            'tanggal': tgl.isoformat() if tgl else '',
            'ref_no': ref,
            'produk': m.product.nama if m.product else '',
            'sku': (m.variant.sku if m.variant and m.variant.sku else (m.product.sku if m.product else '')) or '',
            'awal': _num(m.stok_awal),
            'masuk': qty if is_masuk else 0,
            'keluar': 0 if is_masuk else qty,
            'sisa': _num(m.stok_akhir),
            'subtotal': qty * _num(m.harga_beli),
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'jumlah_mutasi': len(rows),
            'total_masuk': sum(r['masuk'] for r in rows),
            'total_keluar': sum(r['keluar'] for r in rows),
            'total_nilai': sum(r['subtotal'] for r in rows),
        }]},
    }


MASUK_TIPE = {'masuk', 'produksi', 'pengembalian'}
KELUAR_TIPE = {'keluar', 'penjualan'}


def _mv_date(m):
    return m.tanggal or (m.created_at.date() if m.created_at else None)


def _signed_delta(m):
    """Perubahan stok bertanda dari sebuah mutasi."""
    if m.tipe in MASUK_TIPE:
        return _num(m.qty)
    if m.tipe in KELUAR_TIPE:
        return -_num(m.qty)
    # opname: pakai selisih hasil hitung fisik
    return _num(m.stok_akhir) - _num(m.stok_awal)


def _qty_as_of(as_of):
    """Stok per (product_id, variant_id) pada akhir tanggal `as_of`.

    Direkonstruksi mundur: stok sekarang dikurangi seluruh mutasi SETELAH tanggal
    tersebut. Ini satu-satunya cara karena sistem tidak menyimpan snapshot stok.
    """
    cur = {}
    for s in _sku_rows():
        cur[(s['product'].id, s['variant'].id if s['variant'] else None)] = s['qty']
    for m in ProductStockMovement.objects.all().only(
        'product_id', 'variant_id', 'tipe', 'qty', 'stok_awal', 'stok_akhir', 'tanggal', 'created_at'
    ):
        d = _mv_date(m)
        if d and d > as_of:
            k = (m.product_id, m.variant_id)
            if k in cur:
                cur[k] -= _signed_delta(m)
    return cur


def _movements_in_range(params):
    qs = ProductStockMovement.objects.select_related('product', 'variant').all()
    rows = []
    for m in qs:
        d = _mv_date(m)
        if params['start'] and (not d or d < params['start']):
            continue
        if params['end'] and (not d or d > params['end']):
            continue
        rows.append(m)
    return rows


def _sold_qty_map(params):
    """(product_id, variant_id) -> qty keluar (penjualan/keluar) dalam rentang."""
    out = {}
    for m in _movements_in_range(params):
        if m.tipe in KELUAR_TIPE:
            k = (m.product_id, m.variant_id)
            out[k] = out.get(k, 0.0) + _num(m.qty)
    return out


@report('qty-stok-tanggal', 'Qty stok berdasarkan tanggal', [
    ('nama_produk', 'Nama Produk', 'text'), ('brand', 'Brand', 'text'),
    ('grup', 'Grup', 'text'), ('qty_stok', 'Qty Stok', 'qty'),
    ('harga_beli', 'Harga Beli', 'money'), ('ols_harga_jual', 'OLS - Harga Jual', 'money'),
    ('pos_harga_jual', 'POS - Harga Jual', 'money'), ('total', 'Total', 'money'),
])
def rpt_qty_stok_tanggal(params):
    as_of = params['end'] or params['start'] or timezone.now().date()
    qty_map = _qty_as_of(as_of)
    rows = []
    t_qty = t_beli = t_ols = t_pos = 0.0
    for s in _sku_rows():
        key = (s['product'].id, s['variant'].id if s['variant'] else None)
        qty = qty_map.get(key, 0.0)
        nama = f"{s['nama']} - {s['varian']}" if s['varian'] else s['nama']
        if params['search'] and params['search'].lower() not in nama.lower():
            continue
        total = qty * s['harga_beli']
        rows.append({
            'nama_produk': nama, 'brand': s['brand'], 'grup': s['kategori'],
            'qty_stok': qty, 'harga_beli': s['harga_beli'],
            'ols_harga_jual': s['jual_online'], 'pos_harga_jual': s['jual_toko'],
            'total': total,
        })
        t_qty += qty
        t_beli += total
        t_ols += qty * s['jual_online']
        t_pos += qty * s['jual_toko']
    return {
        'rows': rows,
        'summary': {'rows': [{
            'qty_stok': t_qty, 'harga_beli': t_beli,
            'ols_harga_jual': t_ols, 'pos_harga_jual': t_pos,
        }]},
    }


@report('qty-stok-syncron', 'Qty stok per tanggal manual syncron', [
    ('nama_produk', 'Nama Produk', 'text'), ('brand', 'Brand', 'text'),
    ('grup', 'Grup', 'text'), ('qty_stok', 'Qty Stok', 'qty'),
    ('harga_beli', 'Harga Beli', 'money'), ('ols_harga_jual', 'OLS - Harga Jual', 'money'),
    ('pos_harga_jual', 'POS - Harga Jual', 'money'), ('total', 'Total', 'money'),
])
def rpt_qty_stok_syncron(params):
    """Hasil hitung fisik terakhir (Stok Opname) per produk sampai tanggal terpilih."""
    as_of = params['end'] or params['start'] or timezone.now().date()
    qs = (StockOpnameDocumentItem.objects
          .select_related('document', 'product', 'variant', 'product__kategori', 'product__brand')
          .filter(document__status='selesai', document__tanggal__lte=as_of)
          .order_by('document__tanggal', 'id'))
    latest = {}
    for it in qs:
        latest[(it.product_id, it.variant_id)] = it
    rows = []
    for it in latest.values():
        p = it.product
        if not p:
            continue
        nama = f"{p.nama} - {it.variant.nama_varian}" if it.variant else p.nama
        if params['search'] and params['search'].lower() not in nama.lower():
            continue
        qty = _num(it.stok_aktual)
        harga_beli = _num(it.variant.harga_beli if it.variant else p.harga_beli)
        rows.append({
            'nama_produk': nama,
            'brand': p.brand.nama if p.brand else '',
            'grup': p.kategori.nama if p.kategori else 'Umum',
            'qty_stok': qty, 'harga_beli': harga_beli,
            'ols_harga_jual': _num(p.harga_jual_online),
            'pos_harga_jual': _num(p.harga_jual_toko),
            'total': qty * harga_beli,
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'jumlah_produk': len(rows),
            'total_qty_stok': sum(_num(r['qty_stok']) for r in rows),
            'total_nilai': sum(_num(r['total']) for r in rows),
        }]},
    }


@report('tidak-laku', 'Produk Tidak Laku', [
    ('produk', 'Produk', 'text'), ('varian', 'Varian', 'text'),
    ('kategori', 'Kategori', 'text'), ('qty_stok', 'Qty Stok', 'qty'),
    ('harga_beli', 'Harga Beli', 'money'),
    ('harga_jual_online', 'Harga Jual Online', 'money'),
    ('harga_jual_toko', 'Harga Jual di Toko', 'money'),
])
def rpt_tidak_laku(params):
    sold = _sold_qty_map(params)
    rows = []
    for s in _sku_rows():
        key = (s['product'].id, s['variant'].id if s['variant'] else None)
        if sold.get(key, 0.0) > 0:
            continue
        rows.append({
            'produk': s['nama'], 'varian': s['varian'], 'kategori': s['kategori'],
            'qty_stok': s['qty'], 'harga_beli': s['harga_beli'],
            'harga_jual_online': s['jual_online'], 'harga_jual_toko': s['jual_toko'],
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'produk_tidak_laku': len(rows),
            'total_qty_stok': sum(_num(r['qty_stok']) for r in rows),
            'nilai_stok_mengendap': sum(_num(r['qty_stok']) * _num(r['harga_beli']) for r in rows),
        }]},
    }


@report('sisa-grup', 'Sisa Stok Grup Produk', [
    ('kategori', 'Kategori', 'text'), ('qty_stok', 'Qty Stok', 'qty'),
    ('total_jual', 'Total Penjualan', 'money'),
])
def rpt_sisa_grup(params):
    agg = {}
    for s in _sku_rows():
        g = agg.setdefault(s['kategori'], {'qty': 0.0, 'nilai': 0.0})
        g['qty'] += s['qty']
        g['nilai'] += s['qty'] * s['jual_toko']
    rows = [{'kategori': k, 'qty_stok': v['qty'], 'total_jual': v['nilai']}
            for k, v in sorted(agg.items())]
    return {
        'rows': rows,
        'summary': {'rows': [{
            'qty_stok': sum(v['qty'] for v in agg.values()),
            'total': sum(v['nilai'] for v in agg.values()),
        }]},
    }


@report('qty-keluar', 'Qty Produk Keluar', [
    ('produk', 'Produk', 'text'), ('variant', 'Variant', 'text'),
    ('kategori', 'Kategori', 'text'), ('qty', 'Qty', 'qty'),
    ('harga', 'Harga', 'money'), ('subtotal', 'Subtotal', 'money'),
])
def rpt_qty_keluar(params):
    keluar = _sold_qty_map(params)
    rows = []
    tot_qty = tot_nilai = 0.0
    for s in _sku_rows():
        key = (s['product'].id, s['variant'].id if s['variant'] else None)
        qty = keluar.get(key, 0.0)
        if qty <= 0:
            continue
        subtotal = qty * s['harga_beli']
        rows.append({
            'produk': s['nama'], 'variant': s['varian'], 'kategori': s['kategori'],
            'qty': qty, 'harga': s['harga_beli'], 'subtotal': subtotal,
        })
        tot_qty += qty
        tot_nilai += subtotal
    return {
        'rows': rows,
        'summary': {'rows': [{'qty_jual': tot_qty, 'total_jual': tot_nilai}]},
    }


@report('usia-stok', 'Usia Stok Produk', [
    ('produk', 'Produk', 'text'), ('variant', 'Variant', 'text'),
    ('kategori', 'Kategori', 'text'), ('total_sales_qty', 'Total Sales Qty', 'qty'),
    ('avg_sales_qty', 'Avg Sales Qty', 'qty'), ('qty_stok', 'Qty Stok', 'qty'),
    ('stock_age', 'Stock Age (hari)', 'qty'),
])
def rpt_usia_stok(params):
    """Usia stok = jumlah hari sejak mutasi 'masuk' terakhir (aproksimasi —
    sistem tidak melacak batch/lapisan stok)."""
    sold = _sold_qty_map(params)
    today = timezone.now().date()
    last_in = {}
    for m in ProductStockMovement.objects.filter(tipe__in=MASUK_TIPE).only(
        'product_id', 'variant_id', 'tanggal', 'created_at'
    ):
        d = _mv_date(m)
        k = (m.product_id, m.variant_id)
        if d and (k not in last_in or d > last_in[k]):
            last_in[k] = d

    span = 1
    if params['start'] and params['end']:
        span = max(1, (params['end'] - params['start']).days + 1)

    rows = []
    tot_qty = tot_sales = 0.0
    for s in _sku_rows():
        key = (s['product'].id, s['variant'].id if s['variant'] else None)
        sales = sold.get(key, 0.0)
        d = last_in.get(key)
        rows.append({
            'produk': s['nama'], 'variant': s['varian'], 'kategori': s['kategori'],
            'total_sales_qty': sales,
            'avg_sales_qty': round(sales / span, 2),
            'qty_stok': s['qty'],
            'stock_age': (today - d).days if d else '',
        })
        tot_qty += s['qty']
        tot_sales += sales
    return {
        'rows': rows,
        'summary': {'rows': [{'qty_stok': tot_qty, 'total_sales_qty': tot_sales}]},
    }


@report('tingkatan-harga', 'Tingkatan Harga', [
    ('produk', 'Produk', 'text'), ('sku', 'SKU', 'text'),
    ('harga_beli', 'Harga Beli', 'money'), ('tipe_pelanggan', 'Tipe Pelanggan', 'text'),
    ('qty_mulai', 'Qty Mulai', 'qty'), ('harga_jual', 'Harga Jual', 'money'),
])
def rpt_tingkatan_harga(params):
    rows = []
    for p in Product.objects.all().order_by('nama'):
        tiers = p.tiers if isinstance(p.tiers, list) else []
        if not tiers:
            continue
        if params['search'] and params['search'].lower() not in f"{p.nama} {p.sku or ''}".lower():
            continue
        for t in tiers:
            if not isinstance(t, dict):
                continue
            rows.append({
                'produk': p.nama, 'sku': p.sku or '',
                'harga_beli': _num(p.harga_beli),
                'tipe_pelanggan': 'Umum',
                'qty_mulai': _num(t.get('min_qty')),
                'harga_jual': _num(t.get('price')),
            })
    return {'rows': rows}


@report('value-pergerakan', 'Value Pergerakan berdasarkan tanggal', [
    ('tanggal', 'Tanggal', 'date'), ('masuk', 'Masuk', 'money'), ('keluar', 'Keluar', 'money'),
])
def rpt_value_pergerakan(params):
    agg = {}
    for m in _movements_in_range(params):
        d = _mv_date(m)
        if not d:
            continue
        harga = _num(m.harga_beli) or (_num(m.product.harga_beli) if m.product else 0)
        nilai = _num(m.qty) * harga
        row = agg.setdefault(d, {'masuk': 0.0, 'keluar': 0.0})
        if m.tipe in MASUK_TIPE:
            row['masuk'] += nilai
        elif m.tipe in KELUAR_TIPE:
            row['keluar'] += nilai
    rows = [{'tanggal': d.isoformat(), 'masuk': v['masuk'], 'keluar': v['keluar']}
            for d, v in sorted(agg.items(), reverse=True)]
    return {
        'rows': rows,
        'summary': {'rows': [{
            'total_nilai_masuk': sum(r['masuk'] for r in rows),
            'total_nilai_keluar': sum(r['keluar'] for r in rows),
        }]},
    }


# ---------------------------------------------------------------------------
# Penjualan — helper penggabung Order (pesanan advertising) + POSSale (retail)
# ---------------------------------------------------------------------------
def _orders_in_range(params):
    qs = Order.objects.exclude(status_global='batal').prefetch_related('items__product__brand',
                                                                      'items__product__kategori',
                                                                      'items__product__koleksi',
                                                                      'items__variant')
    if params['start']:
        qs = qs.filter(waktu__date__gte=params['start'])
    if params['end']:
        qs = qs.filter(waktu__date__lte=params['end'])
    return qs


def _pos_sales_in_range(params):
    qs = POSSale.objects.filter(status='paid').select_related('kasir', 'pelanggan').prefetch_related(
        'items__product__brand', 'items__product__kategori', 'items__product__koleksi', 'items__variant')
    if params['start']:
        qs = qs.filter(created_at__date__gte=params['start'])
    if params['end']:
        qs = qs.filter(created_at__date__lte=params['end'])
    return qs


def _sale_lines(params):
    """Baris penjualan gabungan (Order + POS) yang dinormalkan.

    Kolom `product` hanya terisi bila item tertaut ke master Produk. Untuk
    OrderItem, tautan itu baru ada sejak FK OrderItem.product ditambahkan —
    baris lama tanpa tautan tetap muncul, tapi tanpa brand/kategori/SKU.
    """
    lines = []
    for o in _orders_in_range(params):
        tgl = o.waktu.date() if o.waktu else None
        for it in o.items.all():
            qty = _num(it.qty)
            total = _num(it.harga_jual)
            dp = _num(o.diskon_persen)
            lines.append({
                'product': it.product, 'variant': it.variant,
                'nama': it.product.nama if it.product else (it.jenis_produk or ''),
                'qty': qty, 'total': total, 'modal': _num(it.biaya_bahan),
                'tanggal': tgl, 'no_pesanan': str(o.id),
                'pelanggan': o.nama or '', 'id_pelanggan': o.nomor_wa or '',
                'oleh': '', 'sumber': o.get_sumber_display() if hasattr(o, 'get_sumber_display') else (o.sumber or ''),
                'pembayaran': o.metode_pembayaran or '',
                'diskon': _num(o.total_harga) * dp / 100.0,
                # --- kolom detail tambahan untuk export lengkap (item-brand) ---
                'waktu': o.waktu,
                'harga_satuan': (total / qty if qty else total),
                'diskon_persen': dp, 'diskon_baris': total * dp / 100.0,
                'pajak_baris': 0.0,  # order advertising tidak menyimpan pajak per baris
                'uom': (it.product.satuan if it.product else ''),
                'qty_konversi': qty,  # OrderItem tidak punya konversi satuan
                'dilayani_oleh': '',
            })
    for s in _pos_sales_in_range(params):
        tgl = s.created_at.date() if s.created_at else None
        s_sub = _num(s.subtotal)
        for it in s.items.all():
            qty = _num(it.qty)
            line_total = _num(it.subtotal)
            modal = _num(it.product.harga_beli) * qty if it.product else 0.0
            # Diskon & pajak POS tersimpan di tingkat nota; sebar proporsional ke baris.
            share = (line_total / s_sub) if s_sub else 0.0
            diskon_baris = _num(s.diskon) * share
            pajak_baris = _num(s.pajak) * share
            dp = (_num(s.diskon) / s_sub * 100.0) if s_sub else 0.0
            lines.append({
                'product': it.product, 'variant': it.variant,
                'nama': it.nama_snapshot or (it.product.nama if it.product else ''),
                'qty': qty, 'total': line_total, 'modal': modal,
                'tanggal': tgl, 'no_pesanan': s.nomor,
                'pelanggan': s.pelanggan.nama if s.pelanggan else '',
                'id_pelanggan': s.pelanggan.nomor_wa if s.pelanggan else '',
                'oleh': s.kasir.username if s.kasir else '',
                'sumber': 'POS', 'pembayaran': s.metode_bayar or '',
                'diskon': 0.0,
                # --- kolom detail tambahan untuk export lengkap (item-brand) ---
                'waktu': s.created_at,
                'harga_satuan': _num(it.harga_snapshot),
                'diskon_persen': dp, 'diskon_baris': diskon_baris,
                'pajak_baris': pajak_baris,
                'uom': (it.uom_kode or (it.product.satuan if it.product else '')),
                'qty_konversi': (_num(it.uom_qty) if it.uom_qty is not None else qty),
                'dilayani_oleh': s.kasir.username if s.kasir else '',
            })
    return lines


def _group_sales(params, keyfn, labelfn):
    """Agregasi baris penjualan menjadi (label, qty, total, modal)."""
    agg = {}
    for ln in _sale_lines(params):
        k = keyfn(ln)
        if k is None:
            continue
        g = agg.setdefault(k, {'label': labelfn(ln), 'qty': 0.0, 'total': 0.0, 'modal': 0.0})
        g['qty'] += ln['qty']
        g['total'] += ln['total']
        g['modal'] += ln['modal']
    return agg


@report('sku', 'Penjualan berdasarkan SKU', [
    ('produk', 'Produk', 'text'), ('nama_alt', 'Nama Produk Alternatif', 'text'),
    ('variant', 'Variant', 'text'), ('grup', 'Grup', 'text'), ('sku', 'SKU', 'text'),
    ('barcode', 'Barcode', 'text'), ('qty_jual', 'Qty Terjual', 'qty'),
    ('total_jual', 'Total Penjualan', 'money'), ('laba', 'Laba', 'money'),
])
def rpt_sku(params):
    agg = {}
    for ln in _sale_lines(params):
        p, v = ln['product'], ln['variant']
        k = (p.id if p else f"txt:{ln['nama']}", v.id if v else None)
        g = agg.setdefault(k, {'ln': ln, 'qty': 0.0, 'total': 0.0, 'modal': 0.0})
        g['qty'] += ln['qty']
        g['total'] += ln['total']
        g['modal'] += ln['modal']
    rows = []
    for g in agg.values():
        p, v = g['ln']['product'], g['ln']['variant']
        rows.append({
            'produk': p.nama if p else g['ln']['nama'],
            'nama_alt': (p.nama_alternatif or '') if p else '',
            'variant': v.nama_varian if v else '',
            'grup': (p.kategori.nama if p and p.kategori else ''),
            'sku': (v.sku if v and v.sku else (p.sku if p else '')) or '',
            'barcode': (p.barcode or '') if p else '',
            'qty_jual': g['qty'], 'total_jual': g['total'],
            'laba': g['total'] - g['modal'],
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'total_qty_terjual': sum(r['qty_jual'] for r in rows),
            'total_penjualan': sum(r['total_jual'] for r in rows),
            'total_laba': sum(r['laba'] for r in rows),
        }]},
    }


@report('kategori', 'Penjualan berdasarkan Kategori', [
    ('kategori', 'Kategori', 'text'), ('qty_jual', 'Qty Terjual', 'qty'),
    ('total_jual', 'Total Penjualan', 'money'),
])
def rpt_kategori(params):
    agg = _group_sales(
        params,
        lambda ln: ln['product'].kategori_id if ln['product'] and ln['product'].kategori_id else 'lain',
        lambda ln: ln['product'].kategori.nama if ln['product'] and ln['product'].kategori else 'Tanpa Kategori',
    )
    rows = [{'kategori': g['label'], 'qty_jual': g['qty'], 'total_jual': g['total']}
            for g in agg.values()]
    return {
        'rows': rows,
        'summary': {'rows': [{
            'total_qty_terjual': sum(r['qty_jual'] for r in rows),
            'total_penjualan': sum(r['total_jual'] for r in rows),
        }]},
    }


@report('brand', 'Penjualan produk berdasarkan Brand', [
    ('brand', 'Brand', 'text'), ('qty_jual', 'Qty Terjual', 'qty'),
    ('total_jual', 'Total Penjualan', 'money'), ('komisi', 'Komisi', 'money'),
    ('laba', 'Laba', 'money'),
])
def rpt_brand(params):
    agg = {}
    for ln in _sale_lines(params):
        p = ln['product']
        k = p.brand_id if p and p.brand_id else 'lain'
        g = agg.setdefault(k, {
            'label': p.brand.nama if p and p.brand else 'Tanpa Brand',
            'persen': _num(p.brand.komisi_persen) if p and p.brand else 0.0,
            'qty': 0.0, 'total': 0.0, 'modal': 0.0,
        })
        g['qty'] += ln['qty']
        g['total'] += ln['total']
        g['modal'] += ln['modal']
    rows = [{
        'brand': g['label'], 'qty_jual': g['qty'], 'total_jual': g['total'],
        'komisi': g['total'] * g['persen'] / 100.0,
        'laba': g['total'] - g['modal'],
    } for g in agg.values()]
    return {
        'rows': rows,
        'summary': {'rows': [{
            'total_qty_terjual': sum(r['qty_jual'] for r in rows),
            'total_penjualan': sum(r['total_jual'] for r in rows),
            'total_komisi': sum(r['komisi'] for r in rows),
            'total_laba': sum(r['laba'] for r in rows),
        }]},
    }


@report('koleksi', 'Penjualan berdasarkan Koleksi', [
    ('koleksi', 'Koleksi', 'text'), ('qty_jual', 'Qty Terjual', 'qty'),
    ('total_jual', 'Total Penjualan', 'money'),
])
def rpt_koleksi(params):
    agg = _group_sales(
        params,
        lambda ln: ln['product'].koleksi_id if ln['product'] and ln['product'].koleksi_id else 'lain',
        lambda ln: ln['product'].koleksi.nama if ln['product'] and ln['product'].koleksi else 'Tanpa Koleksi',
    )
    rows = [{'koleksi': g['label'], 'qty_jual': g['qty'], 'total_jual': g['total']}
            for g in agg.values()]
    return {
        'rows': rows,
        'summary': {'rows': [{
            'total_qty_terjual': sum(r['qty_jual'] for r in rows),
            'total_penjualan': sum(r['total_jual'] for r in rows),
        }]},
    }


@report('qty-terjual', 'Qty Produk Terjual', [
    ('produk', 'Produk', 'text'), ('kategori', 'Kategori', 'text'),
    ('qty_jual', 'Qty Terjual', 'qty'), ('harga_beli', 'Harga Beli', 'money'),
    ('subtotal', 'Subtotal', 'money'),
    ('harga_jual_online', 'Harga Jual Online', 'money'),
    ('harga_jual_toko', 'Harga Jual di Toko', 'money'),
])
def rpt_qty_terjual(params):
    agg = {}
    for ln in _sale_lines(params):
        p = ln['product']
        k = p.id if p else f"txt:{ln['nama']}"
        g = agg.setdefault(k, {'ln': ln, 'qty': 0.0, 'total': 0.0})
        g['qty'] += ln['qty']
        g['total'] += ln['total']
    rows = []
    t_qty = t_beli = t_ols = t_pos = 0.0
    for g in agg.values():
        p = g['ln']['product']
        harga_beli = _num(p.harga_beli) if p else 0.0
        subtotal = g['qty'] * harga_beli
        ols = _num(p.harga_jual_online) if p else 0.0
        pos = _num(p.harga_jual_toko) if p else 0.0
        rows.append({
            'produk': p.nama if p else g['ln']['nama'],
            'kategori': (p.kategori.nama if p and p.kategori else ''),
            'qty_jual': g['qty'], 'harga_beli': harga_beli, 'subtotal': subtotal,
            'harga_jual_online': ols, 'harga_jual_toko': pos,
        })
        t_qty += g['qty']
        t_beli += subtotal
        t_ols += g['qty'] * ols
        t_pos += g['qty'] * pos
    return {
        'rows': rows,
        'summary': {'rows': [{
            'qty_jual': t_qty, 'total_harga_beli': t_beli,
            'total_jual_online': t_ols, 'total_jual_toko': t_pos,
        }]},
    }


@report('material', 'Produk Material', [
    ('produk', 'Produk', 'text'), ('sku', 'SKU', 'text'),
    ('bahan', 'Bahan / Resep', 'text'), ('qty', 'Qty', 'qty'),
    ('unit', 'Unit Pengukuran', 'text'),
    ('nilai_konversi', 'Nilai Konversi - Unit Pengukuran', 'qty'),
    ('harga_beli', 'Harga Beli', 'money'),
    ('harga_beli_material', 'Harga Beli Material', 'money'),
])
def rpt_material(params):
    """Dari field `Product.material`. Resep BoM lama menunjuk model legacy
    (ProductPrice/InventoryItem) sehingga sengaja tidak dipakai di sini."""
    rows = []
    for p in Product.objects.exclude(material='').exclude(material__isnull=True).order_by('nama'):
        if params['search'] and params['search'].lower() not in f"{p.nama} {p.material}".lower():
            continue
        rows.append({
            'produk': p.nama, 'sku': p.sku or '', 'bahan': p.material or '',
            'qty': _num(p.qty_stok), 'unit': p.satuan or '', 'nilai_konversi': 1,
            'harga_beli': _num(p.harga_beli), 'harga_beli_material': _num(p.harga_beli),
        })
    return {'rows': rows}


# --- Laporan Penjualan -----------------------------------------------------
@report('rincian-penjualan', 'Rincian Penjualan', [
    ('no_pesanan', 'No. Pesanan', 'text'), ('tanggal', 'Tanggal', 'date'),
    ('pembayaran', 'Pembayaran', 'text'), ('penjualan_oleh', 'Penjualan Oleh', 'text'),
    ('pelanggan', 'Pelanggan', 'text'), ('id_pelanggan', 'ID Pelanggan', 'text'),
    ('pembulatan', 'Pembulatan', 'money'), ('total_penjualan', 'Total Penjualan', 'money'),
    ('jumlah_ditebus', 'Jumlah Ditebus', 'money'), ('pengiriman_pajak', 'Pengiriman/Pajak', 'money'),
    ('modal_produk', 'Modal Produk', 'money'), ('laba', 'Laba', 'money'),
    ('biaya_layanan', 'Biaya Layanan', 'money'), ('tambahan_pembayaran', 'Tambahan Pembayaran', 'money'),
    ('diskon', 'Diskon', 'money'), ('tebus_deposit', 'Tebus Deposit', 'money'),
    ('pengunjung', 'Pengunjung', 'qty'),
    ('sumber', 'Sumber Pesanan', 'text'), ('waktu', 'Waktu Pesanan', 'text'),
    ('jumlah_item', 'Jumlah Item', 'qty'), ('sisa_tagihan', 'Sisa Tagihan', 'money'),
    ('status', 'Status', 'text'),
])
def rpt_rincian_penjualan(params):
    """Satu baris per nota, gabungan Order + POS.

    Kolom yang memang tidak disimpan sistem (pembulatan, biaya layanan, deposit,
    pengunjung) sengaja diisi 0 — bukan dikarang.
    """
    rows = []
    t_total = t_modal = t_diskon = t_pajak = 0.0
    for o in _orders_in_range(params):
        modal = sum(_num(i.biaya_bahan) for i in o.items.all())
        total = _num(o.total_harga)
        diskon = total * _num(o.diskon_persen) / 100.0
        rows.append({
            'no_pesanan': str(o.id),
            'tanggal': o.waktu.date().isoformat() if o.waktu else '',
            'pembayaran': o.metode_pembayaran or '', 'penjualan_oleh': '',
            'pelanggan': o.nama or '', 'id_pelanggan': o.nomor_wa or '',
            'pembulatan': 0, 'total_penjualan': total, 'jumlah_ditebus': _num(o.dp_dibayar),
            'pengiriman_pajak': 0, 'modal_produk': modal, 'laba': total - modal,
            'biaya_layanan': 0, 'tambahan_pembayaran': 0, 'diskon': diskon,
            'tebus_deposit': 0, 'pengunjung': 0,
            'sumber': (o.get_sumber_display() if hasattr(o, 'get_sumber_display') else (o.sumber or '')),
            'waktu': o.waktu.strftime('%Y-%m-%d %H:%M') if o.waktu else '',
            'jumlah_item': sum(_num(i.qty) for i in o.items.all()),
            'sisa_tagihan': _num(o.sisa_tagihan),
            'status': (o.status_global or ''),
        })
        t_total += total
        t_modal += modal
        t_diskon += diskon
    for s in _pos_sales_in_range(params):
        modal = sum((_num(i.product.harga_beli) * _num(i.qty)) if i.product else 0 for i in s.items.all())
        total = _num(s.total)
        rows.append({
            'no_pesanan': s.nomor,
            'tanggal': s.created_at.date().isoformat() if s.created_at else '',
            'pembayaran': s.metode_bayar or '',
            'penjualan_oleh': s.kasir.username if s.kasir else '',
            'pelanggan': s.pelanggan.nama if s.pelanggan else '',
            'id_pelanggan': s.pelanggan.nomor_wa if s.pelanggan else '',
            'pembulatan': 0, 'total_penjualan': total, 'jumlah_ditebus': _num(s.dibayar),
            'pengiriman_pajak': _num(s.pajak), 'modal_produk': modal, 'laba': total - modal,
            'biaya_layanan': 0, 'tambahan_pembayaran': 0, 'diskon': _num(s.diskon),
            'tebus_deposit': 0, 'pengunjung': 0,
            'sumber': 'POS',
            'waktu': s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else '',
            'jumlah_item': sum(_num(i.qty) for i in s.items.all()),
            'sisa_tagihan': max(_num(s.total) - _num(s.dibayar), 0.0),
            'status': (s.status or ''),
        })
        t_total += total
        t_modal += modal
        t_diskon += _num(s.diskon)
        t_pajak += _num(s.pajak)
    rows.sort(key=lambda r: r['tanggal'], reverse=True)
    return {
        'rows': rows,
        'summary': {'rows': [{
            'mata_uang': 'IDR', 'total_pembulatan': 0, 'total_penjualan': t_total,
            'pengiriman': 0, 'pajak': t_pajak, 'modal_produk': t_modal,
            'laba': t_total - t_modal, 'biaya_layanan': 0, 'tambahan_pembayaran': 0,
            'diskon': t_diskon, 'total_tebus_deposit': 0, 'total_pengunjung': 0,
        }]},
    }


@report('item-penjualan-tanggal', 'Item Penjualan berdasarkan Tanggal', [
    ('no_pesanan', 'No. Pesanan', 'text'), ('waktu', 'Waktu Pesanan', 'text'),
    ('tanggal', 'Tanggal', 'date'), ('sumber', 'Sumber Pesanan', 'text'),
    ('penjualan_oleh', 'Penjualan Oleh', 'text'), ('dilayani_oleh', 'Dilayani Oleh', 'text'),
    ('brand', 'Brand', 'text'), ('grup_item', 'Grup Item', 'text'),
    ('item', 'Item', 'text'), ('sku_item', 'SKU Item', 'text'),
    ('no_seri', 'No. Seri', 'text'), ('pelanggan', 'Pelanggan', 'text'),
    ('id_pelanggan', 'ID Pelanggan', 'text'), ('qty', 'Qty', 'qty'),
    ('uom', 'Satuan', 'text'), ('qty_konversi', 'Qty (Satuan Terpilih)', 'qty'),
    ('mata_uang', 'Mata Uang', 'text'), ('harga', 'Harga Satuan', 'money'),
    ('diskon_persen', 'Diskon (%)', 'qty'), ('diskon', 'Nilai Diskon', 'money'),
    ('pajak', 'Pajak', 'money'), ('total_penjualan', 'Total Penjualan', 'money'),
    ('modal_satuan', 'Modal / Unit', 'money'), ('modal_produk', 'Modal Produk', 'money'),
    ('laba', 'Laba', 'money'), ('komisi', 'Komisi', 'money'),
    ('cara_pembayaran', 'Cara Pembayaran', 'text'),
])
def rpt_item_penjualan(params):
    """Rincian penjualan per-item lengkap (Order + POS) dengan atribut waktu,
    sumber, brand, grup, SKU, satuan, harga, diskon, pajak, modal, laba, dan
    komisi — supaya export memuat data utuh, bukan hanya sebagian kolom."""
    rows = []
    t_qty = t_diskon = t_total = t_modal = 0.0
    for ln in sorted(_sale_lines(params), key=lambda x: (x['tanggal'] or timezone.now().date()), reverse=True):
        p, v = ln['product'], ln['variant']
        brand = p.brand if p and p.brand else None
        komisi = 0.0
        if brand and _num(brand.komisi_persen):
            komisi = ln['total'] * _num(brand.komisi_persen) / 100.0
        qty = ln['qty']
        modal = ln['modal']
        amount = ln['total']
        diskon_baris = ln.get('diskon_baris', 0.0)
        w = ln.get('waktu')
        rows.append({
            'no_pesanan': ln['no_pesanan'],
            'waktu': w.strftime('%Y-%m-%d %H:%M') if w else (ln['tanggal'].isoformat() if ln['tanggal'] else ''),
            'tanggal': ln['tanggal'].isoformat() if ln['tanggal'] else '',
            'sumber': ln.get('sumber', ''),
            'penjualan_oleh': ln.get('oleh', ''),
            'dilayani_oleh': ln.get('dilayani_oleh', ''),
            'brand': brand.nama if brand else 'Tanpa Brand',
            'grup_item': (p.kategori.nama if p and p.kategori else ''),
            'item': ln['nama'],
            'sku_item': ((v.sku if v and v.sku else (p.sku if p else '')) or ''),
            'no_seri': '',
            'pelanggan': ln['pelanggan'],
            'id_pelanggan': ln.get('id_pelanggan', ''),
            'qty': qty, 'uom': ln.get('uom', ''),
            'qty_konversi': ln.get('qty_konversi', qty),
            'mata_uang': 'IDR',
            'harga': ln.get('harga_satuan', (amount / qty if qty else amount)),
            'diskon_persen': ln.get('diskon_persen', 0.0),
            'diskon': diskon_baris,
            'pajak': ln.get('pajak_baris', 0.0),
            'total_penjualan': amount,
            'modal_satuan': (modal / qty if qty else modal),
            'modal_produk': modal,
            'laba': amount - modal, 'komisi': komisi,
            'cara_pembayaran': ln.get('pembayaran', ''),
        })
        t_qty += qty
        t_diskon += diskon_baris
        t_total += amount
        t_modal += modal
    return {
        'rows': rows,
        'summary': {'type': 'grid', 'items': [
            {'label': 'Qty', 'value': t_qty, 'type': 'qty'},
            {'label': 'Diskon Penjualan', 'value': t_diskon, 'type': 'money'},
            {'label': 'Total Penjualan', 'value': t_total, 'type': 'money'},
            {'label': 'Modal Produk', 'value': t_modal, 'type': 'money'},
            {'label': 'Total Laba', 'value': t_total - t_modal, 'type': 'money'},
        ]},
    }


@report('item-brand', 'Item Penjualan Berdasarkan Brand', [
    ('no_pesanan', 'No. Pesanan', 'text'), ('waktu', 'Waktu Pesanan', 'text'),
    ('sumber', 'Sumber Pesanan', 'text'), ('sales_oleh', 'Penjualan Oleh', 'text'),
    ('dilayani_oleh', 'Dilayani Oleh', 'text'), ('brand', 'Brand', 'text'),
    ('komisi_persen', 'Rate Komisi Brand (%)', 'qty'),
    ('komisi_nilai', 'Nilai Komisi Brand', 'money'),
    ('grup_item', 'Grup Item', 'text'), ('nama_item', 'Nama Item', 'text'),
    ('sku_item', 'SKU Item', 'text'), ('pelanggan', 'Pelanggan', 'text'),
    ('qty', 'Qty', 'qty'), ('uom', 'Satuan', 'text'),
    ('qty_konversi', 'Qty (Satuan Terpilih)', 'qty'),
    ('mata_uang', 'Mata Uang', 'text'), ('harga', 'Harga Satuan', 'money'),
    ('harga_addon', 'Harga Add-on', 'money'),
    ('diskon_persen', 'Diskon (%)', 'qty'), ('diskon_nilai', 'Nilai Diskon', 'money'),
    ('jumlah', 'Jumlah', 'money'), ('pajak', 'Pajak', 'money'),
    ('modal_satuan', 'Modal / Unit', 'money'), ('modal_total', 'Total Modal', 'money'),
    ('laba', 'Laba', 'money'), ('dibayar_ke_brand', 'Dibayar ke Brand', 'money'),
    ('cara_pembayaran', 'Cara Pembayaran', 'text'),
])
def rpt_item_brand(params):
    """Rincian penjualan per-item lengkap dengan atribut brand, harga, diskon,
    pajak, modal, laba, dan komisi brand — supaya export Excel memuat data utuh
    (bukan sekadar ringkasan beberapa kolom). Kolom yang belum punya sumber data
    di sistem (mis. harga add-on per baris) sengaja diisi 0/kosong, bukan
    dihilangkan, agar struktur export tetap konsisten."""
    rows = []
    t_total = t_modal = t_komisi = t_diskon = t_pajak = 0.0
    for ln in _sale_lines(params):
        p, v = ln['product'], ln['variant']
        brand = p.brand if p and p.brand else None
        rate = _num(brand.komisi_persen) if brand else 0.0
        amount = ln['total']
        komisi = amount * rate / 100.0
        qty = ln['qty']
        modal = ln['modal']
        w = ln.get('waktu')
        rows.append({
            'no_pesanan': ln['no_pesanan'],
            'waktu': w.strftime('%Y-%m-%d %H:%M') if w else (ln['tanggal'].isoformat() if ln['tanggal'] else ''),
            'sumber': ln.get('sumber', ''),
            'sales_oleh': ln.get('oleh', ''),
            'dilayani_oleh': ln.get('dilayani_oleh', ''),
            'brand': brand.nama if brand else 'Tanpa Brand',
            'komisi_persen': rate, 'komisi_nilai': komisi,
            'grup_item': (p.kategori.nama if p and p.kategori else ''),
            'nama_item': ln['nama'],
            'sku_item': ((v.sku if v and v.sku else (p.sku if p else '')) or ''),
            'pelanggan': ln['pelanggan'],
            'qty': qty, 'uom': ln.get('uom', ''),
            'qty_konversi': ln.get('qty_konversi', qty),
            'mata_uang': 'IDR',
            'harga': ln.get('harga_satuan', (amount / qty if qty else amount)),
            'harga_addon': 0.0,  # belum ada model harga add-on per baris
            'diskon_persen': ln.get('diskon_persen', 0.0),
            'diskon_nilai': ln.get('diskon_baris', 0.0),
            'jumlah': amount, 'pajak': ln.get('pajak_baris', 0.0),
            'modal_satuan': (modal / qty if qty else modal),
            'modal_total': modal, 'laba': amount - modal,
            'dibayar_ke_brand': komisi,
            'cara_pembayaran': ln.get('pembayaran', ''),
        })
        t_total += amount
        t_modal += modal
        t_komisi += komisi
        t_diskon += ln.get('diskon_baris', 0.0)
        t_pajak += ln.get('pajak_baris', 0.0)
    return {
        'rows': rows,
        'summary': {'rows': [{
            'mata_uang': 'IDR', 'total_penjualan': t_total, 'total_diskon': t_diskon,
            'total_pajak': t_pajak, 'modal_produk': t_modal,
            'komisi_brand': t_komisi, 'laba': t_total - t_modal, 'total': t_total,
        }]},
    }


@report('pelunasan-kredit', 'Item Penjualan berdasarkan Pelunasan Kredit', [
    ('no_pesanan', 'No. Pesanan', 'text'), ('tanggal', 'Tanggal', 'date'),
    ('waktu', 'Waktu Pesanan', 'text'), ('sumber', 'Sumber Pesanan', 'text'),
    ('penjualan_oleh', 'Penjualan Oleh', 'text'), ('item', 'Item', 'text'),
    ('pelanggan', 'Pelanggan', 'text'), ('id_pelanggan', 'ID Pelanggan', 'text'),
    ('total_penjualan', 'Total Penjualan', 'money'),
    ('cara_pembayaran', 'Cara Pembayaran', 'text'),
    ('jumlah_terbayar', 'Jumlah Terbayar', 'money'),
    ('sisa_tagihan', 'Sisa Tagihan', 'money'),
])
def rpt_pelunasan_kredit(params):
    """Pesanan kredit = ada pembayaran sebagian (dp_dibayar > 0). Sistem belum
    punya tabel pembayaran penjualan, jadi 'Jumlah Terbayar' memakai akumulator
    Order.dp_dibayar."""
    rows = []
    t_total = t_bayar = t_sisa = 0.0
    for o in _orders_in_range(params):
        dp = _num(o.dp_dibayar)
        if dp <= 0:
            continue
        items = ', '.join(filter(None, [(i.product.nama if i.product else i.jenis_produk) for i in o.items.all()]))
        total = _num(o.total_harga)
        sisa = _num(o.sisa_tagihan)
        rows.append({
            'no_pesanan': str(o.id),
            'tanggal': o.waktu.date().isoformat() if o.waktu else '',
            'waktu': o.waktu.strftime('%Y-%m-%d %H:%M') if o.waktu else '',
            'sumber': (o.get_sumber_display() if hasattr(o, 'get_sumber_display') else (o.sumber or '')),
            'penjualan_oleh': '', 'item': items,
            'pelanggan': o.nama or '', 'id_pelanggan': o.nomor_wa or '',
            'total_penjualan': total,
            'cara_pembayaran': o.metode_pembayaran or '',
            'jumlah_terbayar': dp,
            'sisa_tagihan': sisa,
        })
        t_total += total
        t_bayar += dp
        t_sisa += sisa
    return {
        'rows': rows,
        'summary': {'rows': [{'mata_uang': 'IDR', 'total_penjualan': t_total, 'total_terbayar': t_bayar, 'sisa_tagihan': t_sisa, 'total': t_total}]},
    }


# --- Laporan Pembayaran ----------------------------------------------------
@report('ringkasan-metode', 'Ringkasan Metode Pembayaran', [
    ('cara_pembayaran', 'Cara Pembayaran', 'text'),
    ('akun_bank', 'Akun Bank', 'text'), ('jumlah', 'Jumlah', 'money'),
])
def rpt_ringkasan_metode(params):
    """Kolom 'Akun Bank' selalu kosong: sistem belum punya model rekening bank."""
    agg = {}
    for o in _orders_in_range(params):
        k = (o.metode_pembayaran or 'tunai').lower()
        agg[k] = agg.get(k, 0.0) + _num(o.dp_dibayar)
    for s in _pos_sales_in_range(params):
        k = (s.metode_bayar or 'tunai').lower()
        agg[k] = agg.get(k, 0.0) + _num(s.dibayar)
    rows = [{'cara_pembayaran': k.title(), 'akun_bank': '', 'jumlah': v}
            for k, v in sorted(agg.items())]
    return {
        'rows': rows,
        'summary': {'rows': [{'mata_uang': 'IDR', 'jumlah': sum(agg.values())}]},
    }


@report('pembayaran-sudah-lunas', 'Pembayaran yang sudah lunas', [
    ('no_pesanan', 'No. Pesanan', 'text'), ('tanggal_pembayaran', 'Tanggal Pembayaran', 'date'),
    ('pelanggan', 'Pelanggan', 'text'), ('id_pelanggan', 'ID Pelanggan', 'text'),
    ('cara_pembayaran', 'Cara Pembayaran', 'text'),
    ('sumber_pesanan', 'Sumber Pesanan', 'text'),
    ('total_penjualan', 'Total Penjualan', 'money'),
    ('total_pembayaran', 'Total Pembayaran', 'money'),
])
def rpt_pembayaran_lunas(params):
    rows = []
    t_jual = t_bayar = 0.0
    for o in _orders_in_range(params):
        if _num(o.sisa_tagihan) > 0:
            continue
        total = _num(o.total_harga)
        rows.append({
            'no_pesanan': str(o.id),
            'tanggal_pembayaran': o.waktu.date().isoformat() if o.waktu else '',
            'pelanggan': o.nama or '', 'id_pelanggan': o.nomor_wa or '',
            'cara_pembayaran': o.metode_pembayaran or '',
            'sumber_pesanan': (o.get_sumber_display() if hasattr(o, 'get_sumber_display') else (o.sumber or '')),
            'total_penjualan': total, 'total_pembayaran': _num(o.dp_dibayar),
        })
        t_jual += total
        t_bayar += _num(o.dp_dibayar)
    for s in _pos_sales_in_range(params):
        total = _num(s.total)
        rows.append({
            'no_pesanan': s.nomor,
            'tanggal_pembayaran': s.created_at.date().isoformat() if s.created_at else '',
            'pelanggan': s.pelanggan.nama if s.pelanggan else '',
            'id_pelanggan': s.pelanggan.nomor_wa if s.pelanggan else '',
            'cara_pembayaran': s.metode_bayar or '', 'sumber_pesanan': 'POS',
            'total_penjualan': total, 'total_pembayaran': _num(s.dibayar),
        })
        t_jual += total
        t_bayar += _num(s.dibayar)
    return {
        'rows': rows,
        'summary': {'rows': [{
            'mata_uang': 'IDR', 'total_penjualan': t_jual, 'total_pembayaran': t_bayar,
        }]},
    }


@report('pembayaran-belum-lunas', 'Pembayaran yang belum lunas', [
    ('no_pesanan', 'No. Pesanan', 'text'), ('tanggal', 'Tanggal', 'date'),
    ('waktu', 'Waktu Pesanan', 'text'), ('sumber_pesanan', 'Sumber Pesanan', 'text'),
    ('pelanggan', 'Pelanggan', 'text'), ('id_pelanggan', 'ID Pelanggan', 'text'),
    ('cara_pembayaran', 'Cara Pembayaran', 'text'),
    ('total_penjualan', 'Total Penjualan', 'money'),
    ('telah_dibayar', 'Telah Dibayar', 'money'), ('sisa', 'Sisa Tagihan', 'money'),
])
def rpt_pembayaran_belum_lunas(params):
    rows = []
    t_sisa = 0.0
    for o in _orders_in_range(params):
        sisa = _num(o.sisa_tagihan)
        if sisa <= 0:
            continue
        rows.append({
            'no_pesanan': str(o.id),
            'tanggal': o.waktu.date().isoformat() if o.waktu else '',
            'waktu': o.waktu.strftime('%Y-%m-%d %H:%M') if o.waktu else '',
            'sumber_pesanan': (o.get_sumber_display() if hasattr(o, 'get_sumber_display') else (o.sumber or '')),
            'pelanggan': o.nama or '', 'id_pelanggan': o.nomor_wa or '',
            'cara_pembayaran': o.metode_pembayaran or '',
            'total_penjualan': _num(o.total_harga), 'telah_dibayar': _num(o.dp_dibayar),
            'sisa': sisa,
        })
        t_sisa += sisa
    rows.sort(key=lambda r: r['sisa'], reverse=True)
    return {'rows': rows, 'summary': {'rows': [{'mata_uang': 'IDR', 'jumlah': t_sisa}]}}


@report('penjualan-pembayaran-pelanggan', 'Penjualan berdasarkan Pembayaran Pelanggan', [
    ('sumber_penjualan', 'Sumber Penjualan', 'text'),
    ('tipe_pelanggan', 'Tipe Pelanggan', 'text'), ('pelanggan', 'Pelanggan', 'text'),
    ('total_penjualan', 'Total Penjualan', 'money'),
    ('biaya_pengiriman', 'Biaya Pengiriman', 'money'),
    ('total_minus_pengiriman', 'Total Penjualan - Biaya Pengiriman', 'money'),
])
def rpt_penjualan_pembayaran_pelanggan(params):
    """'Tipe Pelanggan' kosong: pesanan belum tertaut ke master Pelanggan/grup.
    'Biaya Pengiriman' 0: tidak disimpan pada penjualan."""
    agg = {}
    for o in _orders_in_range(params):
        k = ('Pesanan', o.nomor_wa or o.nama or '-')
        g = agg.setdefault(k, {'sumber': o.sumber or 'Pesanan', 'pelanggan': o.nama or '-', 'total': 0.0})
        g['total'] += _num(o.total_harga)
    for s in _pos_sales_in_range(params):
        nm = s.pelanggan.nama if s.pelanggan else '-'
        k = ('POS', nm)
        g = agg.setdefault(k, {'sumber': 'POS', 'pelanggan': nm, 'total': 0.0})
        g['total'] += _num(s.total)
    rows = [{
        'sumber_penjualan': g['sumber'], 'tipe_pelanggan': '', 'pelanggan': g['pelanggan'],
        'total_penjualan': g['total'], 'biaya_pengiriman': 0,
        'total_minus_pengiriman': g['total'],
    } for g in agg.values()]
    total = sum(g['total'] for g in agg.values())
    return {
        'rows': rows,
        'summary': {'rows': [{
            'mata_uang': 'IDR', 'total_penjualan': total, 'biaya_pengiriman': 0,
            'total_minus_pengiriman': total,
        }]},
    }


@report('sisa-uom', 'Sisa stok berdasarkan satuan (UOM)', [
    ('produk', 'Produk', 'text'), ('uom', 'UOM', 'text'),
    ('qty_stok', 'Qty Stok', 'qty'), ('harga_beli', 'Harga Beli', 'money'),
    ('subtotal', 'Subtotal', 'money'),
    ('harga_jual_online', 'Harga Jual Online', 'money'),
    ('harga_jual_toko', 'Harga Jual Toko', 'money'),
])
def rpt_sisa_uom(params):
    """Sisa stok dinyatakan dalam tiap satuan yang dikonfigurasi produk.

    Stok disimpan dalam satuan DASAR; baris satuan alternatif adalah hasil
    konversi (stok dasar ÷ konverter). Baris satuan dasar selalu ikut
    ditampilkan sebagai pembanding.
    """
    from . import uom as uom_helper

    rows = []
    t_sub = 0.0
    for s in _sku_rows():
        p = s['product']
        nama = f"{s['nama']} - {s['varian']}" if s['varian'] else s['nama']
        if params['search'] and params['search'].lower() not in nama.lower():
            continue

        # Satuan dasar
        rows.append({
            'produk': nama, 'uom': p.satuan or 'pcs',
            'qty_stok': s['qty'], 'harga_beli': s['harga_beli'],
            'subtotal': s['qty'] * s['harga_beli'],
            'harga_jual_online': s['jual_online'], 'harga_jual_toko': s['jual_toko'],
        })
        t_sub += s['qty'] * s['harga_beli']

        if not p.uom_enabled:
            continue
        for u in uom_helper.get_units(p, s['variant']):
            k = float(uom_helper.konverter(u))
            if k <= 0:
                continue
            qty_unit = s['qty'] / k
            harga_beli_unit = _num(u.get('harga_beli')) or (s['harga_beli'] * k)
            rows.append({
                'produk': nama,
                'uom': u.get('nama_satuan') or u.get('kode_satuan') or '-',
                'qty_stok': round(qty_unit, 4),
                'harga_beli': harga_beli_unit,
                'subtotal': qty_unit * harga_beli_unit,
                'harga_jual_online': _num(u.get('harga_jual_online')) or (s['jual_online'] * k),
                'harga_jual_toko': _num(u.get('harga_jual_toko')) or (s['jual_toko'] * k),
            })

    return {
        'rows': rows,
        'summary': {'rows': [{
            'jumlah_baris': len(rows),
            'nilai_stok_satuan_dasar': t_sub,
        }]},
    }


# ---------------------------------------------------------------------------
# Laporan FIFO / Kedaluwarsa (lapisan biaya)
# ---------------------------------------------------------------------------
def _layer_rows(params, hanya_kedaluwarsa=False):
    from .product_models import StockLayer
    qs = StockLayer.objects.select_related('product', 'variant').all()
    if hanya_kedaluwarsa:
        qs = qs.filter(tanggal_kadaluwarsa__isnull=False).order_by('tanggal_kadaluwarsa', 'id')
    else:
        qs = qs.order_by('tanggal_masuk', 'id')
    if params['start']:
        qs = qs.filter(tanggal_masuk__gte=params['start'])
    if params['end']:
        qs = qs.filter(tanggal_masuk__lte=params['end'])

    rows = []
    for l in qs:
        nama = f"{l.product.nama} - {l.variant.nama_varian}" if l.variant else l.product.nama
        if params['search'] and params['search'].lower() not in nama.lower():
            continue
        masuk = _num(l.qty_masuk)
        sisa = _num(l.sisa_qty)
        rows.append({
            'layer': l, 'nama': nama, 'masuk': masuk, 'sisa': sisa,
            'keluar': masuk - sisa,
        })
    return rows


@report('stok-fifo', 'Stok FIFO', [
    ('tanggal_masuk', 'Tanggal Masuk', 'date'), ('produk', 'Produk', 'text'),
    ('sku', 'SKU', 'text'), ('masuk', 'Masuk', 'qty'), ('keluar', 'Keluar', 'qty'),
    ('sisa', 'Sisa', 'qty'), ('harga_beli', 'Harga Beli', 'money'), ('rak', 'Rak', 'text'),
])
def rpt_stok_fifo(params):
    rows = []
    for r in _layer_rows(params):
        l = r['layer']
        rows.append({
            'tanggal_masuk': l.tanggal_masuk.isoformat() if l.tanggal_masuk else '',
            'produk': r['nama'],
            'sku': (l.variant.sku if l.variant and l.variant.sku else (l.product.sku or '')),
            'masuk': r['masuk'], 'keluar': r['keluar'], 'sisa': r['sisa'],
            'harga_beli': _num(l.harga_beli), 'rak': l.rak or '',
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'jumlah_lapisan': len(rows),
            'total_masuk': sum(r['masuk'] for r in rows),
            'total_sisa': sum(r['sisa'] for r in rows),
            'nilai_sisa': sum(r['sisa'] * r['harga_beli'] for r in rows),
        }]},
    }


@report('stok-kedaluwarsa', 'Stok Kedaluwarsa', [
    ('tanggal_kedaluwarsa', 'Tanggal Kedaluwarsa', 'date'), ('produk', 'Produk', 'text'),
    ('sku', 'SKU', 'text'), ('masuk', 'Masuk', 'qty'), ('keluar', 'Keluar', 'qty'),
    ('sisa', 'Sisa', 'qty'), ('harga_beli', 'Harga Beli', 'money'), ('rak', 'Rak', 'text'),
])
def rpt_stok_kedaluwarsa(params):
    rows = []
    for r in _layer_rows(params, hanya_kedaluwarsa=True):
        l = r['layer']
        rows.append({
            'tanggal_kedaluwarsa': l.tanggal_kadaluwarsa.isoformat() if l.tanggal_kadaluwarsa else '',
            'produk': r['nama'],
            'sku': (l.variant.sku if l.variant and l.variant.sku else (l.product.sku or '')),
            'masuk': r['masuk'], 'keluar': r['keluar'], 'sisa': r['sisa'],
            'harga_beli': _num(l.harga_beli), 'rak': l.rak or '',
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'batch_berkedaluwarsa': len(rows),
            'total_sisa': sum(r['sisa'] for r in rows),
            'nilai_sisa': sum(r['sisa'] * r['harga_beli'] for r in rows),
        }]},
    }


@report('pergerakan-fifo', 'Pergerakan Stok Kedaluwarsa & FIFO', [
    ('tanggal', 'Tanggal', 'date'), ('produk', 'Produk', 'text'),
    ('no_transaksi', 'No. Transaksi', 'text'), ('awal', 'Awal', 'qty'),
    ('masuk', 'Masuk', 'qty'), ('keluar', 'Keluar', 'qty'), ('sisa', 'Sisa', 'qty'),
])
def rpt_pergerakan_fifo(params):
    """Mutasi stok dengan saldo berjalan, dilihat dari sisi lapisan FIFO."""
    rows = []
    for m in sorted(_movements_in_range(params), key=lambda x: (_mv_date(x) or timezone.now().date(), x.id)):
        nama = f"{m.product.nama} - {m.variant.nama_varian}" if m.variant else m.product.nama
        if params['search'] and params['search'].lower() not in nama.lower():
            continue
        masuk = _num(m.qty) if m.tipe in MASUK_TIPE else 0
        keluar = _num(m.qty) if m.tipe in KELUAR_TIPE else 0
        no_tx = ''
        for atribut in ('stock_in_document', 'stock_out_document',
                        'stock_production_document', 'stock_opname_document'):
            doc = getattr(m, atribut, None)
            if doc is not None:
                no_tx = doc.nomor or ''
                break
        rows.append({
            'tanggal': (_mv_date(m) or '').isoformat() if _mv_date(m) else '',
            'produk': nama, 'no_transaksi': no_tx,
            'awal': _num(m.stok_awal), 'masuk': masuk, 'keluar': keluar,
            'sisa': _num(m.stok_akhir),
        })
    rows.reverse()
    return {
        'rows': rows,
        'summary': {'rows': [{
            'jumlah_mutasi': len(rows),
            'total_masuk': sum(r['masuk'] for r in rows),
            'total_keluar': sum(r['keluar'] for r in rows),
        }]},
    }


@report('banding-fifo', 'Perbandingan Stok & Stok FIFO', [
    ('produk', 'Produk', 'text'), ('stok', 'Stok', 'qty'), ('stok_fifo', 'Stok FIFO', 'qty'),
    ('selisih', 'Selisih', 'qty'),
])
def rpt_banding_fifo(params):
    """Rekonsiliasi: qty_stok produk vs total sisa lapisan FIFO.

    Selisih != 0 menandakan lapisan belum disinkronkan (tekan Sync Stok Produk)
    atau ada mutasi yang lolos dari pencatatan lapisan.
    """
    from . import stock_fifo
    rekon = stock_fifo.reconciliation_report()
    rows = []
    for it in rekon['tidak_cocok']:
        if params['search'] and params['search'].lower() not in it['produk'].lower():
            continue
        rows.append({
            'produk': it['produk'], 'stok': it['qty_stok'],
            'stok_fifo': it['stok_fifo'], 'selisih': it['qty_stok'] - it['stok_fifo'],
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'produk_tidak_cocok': len(rows),
            'total_selisih': sum(r['selisih'] for r in rows),
        }]},
    }


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------
@report('shortfall-stok', 'Shortfall Stok (Konsumsi Melebihi Lapisan)', [
    ('tanggal', 'Tanggal', 'date'), ('produk', 'Produk', 'text'),
    ('sku', 'SKU', 'text'), ('qty', 'Qty Kurang', 'qty'),
    ('harga_beli', 'Harga Beli Fallback', 'money'), ('nilai', 'Nilai Shortfall', 'money'),
])
def rpt_shortfall_stok(params):
    """BE-29: tampilkan akumulasi shortfall FIFO (is_shortfall=True) sebagai laporan/alert.

    Setiap baris shortfall menandakan konsumsi stok melebihi lapisan yang tersedia
    (indikasi stok minus / data drift) yang sebelumnya hanya tercatat di log.
    """
    qs = StockLayerConsumption.objects.filter(is_shortfall=True).select_related('product', 'variant')
    if params.get('start'):
        qs = qs.filter(created_at__date__gte=params['start'])
    if params.get('end'):
        qs = qs.filter(created_at__date__lte=params['end'])
    search = params.get('search')
    if search:
        qs = qs.filter(Q(product__nama__icontains=search))

    rows = []
    total_qty = 0.0
    total_nilai = 0.0
    for c in qs.order_by('-created_at'):
        q = _num(c.qty)
        hb = _num(c.harga_beli)
        nilai = q * hb
        total_qty += q
        total_nilai += nilai
        rows.append({
            'tanggal': c.created_at.date().isoformat() if c.created_at else '',
            'produk': c.product.nama,
            'sku': (c.variant.sku if c.variant and c.variant.sku else (c.product.sku or '')),
            'qty': q,
            'harga_beli': hb,
            'nilai': nilai,
        })
    return {
        'rows': rows,
        'summary': {'rows': [{
            'jumlah_kejadian': len(rows),
            'total_qty_kurang': total_qty,
            'total_nilai_shortfall': total_nilai,
        }]},
    }


class ReportDataView(APIView):
    permission_classes = [IsAuthenticated]
    throttle_classes = [ReportRateThrottle]
    max_rows = 1000

    def get(self, request, report_id):
        entry = REPORT_REGISTRY.get(report_id)
        if not entry:
            return Response(
                {'error': f"Laporan '{report_id}' belum tersedia di backend."},
                status=404,
            )
        result = entry['handler'](_params(request)) or {}
        all_rows = result.get('rows', [])
        return Response({
            'label': entry['label'],
            # Sertakan seluruh kolom yang ada di data (bukan hanya yang
            # dideklarasikan) supaya tampilan & export selalu lengkap.
            'columns': _effective_columns(entry['columns'], all_rows),
            'rows': all_rows[:self.max_rows],
            'truncated': len(all_rows) > self.max_rows,
            'summary': result.get('summary'),
        })


def _humanize(key):
    """'total_penjualan' -> 'Total Penjualan'."""
    return str(key).replace('_', ' ').strip().title()


def _summary_lines(summary):
    """Ratakan ringkasan (bentuk apa pun) jadi daftar (label, nilai) untuk export.

    Mendukung tiga bentuk yang dipakai handler: {'rows': [...]},
    {'type': 'list'|'grid', 'items': [{'label','value'}]}.
    """
    if not summary:
        return []
    lines = []
    if summary.get('items'):
        for it in summary['items']:
            lines.append((it.get('label', ''), it.get('value', '')))
        return lines
    for row in summary.get('rows', []) or []:
        for k, v in row.items():
            lines.append((_humanize(k), v))
    return lines


def _effective_columns(columns, rows):
    """Gabungkan kolom yang dideklarasikan laporan dengan SEMUA key lain yang
    muncul di data baris, sehingga export Excel / PDF / tampilan layar tidak
    pernah membuang kolom yang sebenarnya ada di data. Kolom deklaratif tetap
    di depan (urutan, label, dan tipe aslinya dipertahankan); key tambahan yang
    ditemukan pada baris diappend dengan label yang di-humanize dan tipe 'text'.

    Ini menjadi jaring pengaman menyeluruh: laporan apa pun (sekarang maupun
    yang ditambahkan nanti) otomatis mengekspor seluruh kolom datanya.
    """
    cols = list(columns or [])
    seen = {c['key'] for c in cols}
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                cols.append({'key': k, 'label': _humanize(k), 'type': 'text'})
    return cols


class ReportExportView(APIView):
    """Export laporan: ?format=xlsx (default) atau pdf (HTML siap cetak).

    Memakai IgnoreFormatContentNegotiation karena DRF secara default menafsirkan
    query param `format` sebagai penentu renderer dan akan menolak 'xlsx' (406).
    """
    permission_classes = [IsAuthenticated]
    content_negotiation_class = IgnoreFormatContentNegotiation

    def get(self, request, report_id):
        entry = REPORT_REGISTRY.get(report_id)
        if not entry:
            return Response({'error': f"Laporan '{report_id}' belum tersedia."}, status=404)

        result = entry['handler'](_params(request)) or {}
        rows = result.get('rows', [])
        # Jaring pengaman: ekspor SEMUA kolom yang muncul di data, bukan hanya
        # kolom yang dideklarasikan, agar tidak ada data yang terpotong.
        columns = _effective_columns(entry['columns'], rows)
        ringkasan = _summary_lines(result.get('summary'))
        stamp = timezone.now().strftime('%Y%m%d')
        fmt = (request.query_params.get('format') or 'xlsx').lower()

        if fmt == 'pdf':
            return self._html(entry, columns, rows, ringkasan)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_id}_{stamp}.xlsx"'
        self._write_xlsx(response, entry, columns, rows, ringkasan)
        return response

    def _write_xlsx(self, response, entry, columns, rows, ringkasan):
        """Tulis workbook yang RAPI: header tebal berwarna, lebar kolom otomatis,
        format angka (ribuan / desimal), border, freeze header, dan auto-filter —
        supaya hasil unduhan tidak berdempetan / berantakan."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        MONEY_FMT = '#,##0'
        QTY_FMT = '#,##0.###'
        thin = Side(style='thin', color='D0D7DE')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill('solid', fgColor='1F2937')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (entry['label'][:31] or 'Laporan')
        n_data = len(rows)

        # --- Header ---
        ws.append([c['label'] for c in columns])
        for ci, _c in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[1].height = 28

        # --- Baris data + format sesuai tipe kolom ---
        for r in rows:
            ws.append([r.get(c['key'], '') for c in columns])
        for ci, c in enumerate(columns, start=1):
            ctype = c.get('type')
            for ri in range(2, 2 + n_data):
                cell = ws.cell(row=ri, column=ci)
                cell.border = border
                if ctype == 'money':
                    cell.number_format = MONEY_FMT
                    cell.alignment = Alignment(horizontal='right', vertical='top')
                elif ctype == 'qty':
                    cell.number_format = QTY_FMT
                    cell.alignment = Alignment(horizontal='right', vertical='top')
                elif ctype == 'date':
                    cell.alignment = Alignment(horizontal='center', vertical='top')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

        # --- Lebar kolom otomatis (berdasar konten, dibatasi 12..48) ---
        for ci, c in enumerate(columns, start=1):
            max_len = len(str(c['label']))
            for r in rows:
                v = r.get(c['key'], '')
                if isinstance(v, bool):
                    s = str(v)
                elif isinstance(v, (int, float)):
                    s = f'{v:,.0f}' if c.get('type') == 'money' else f'{v:,.3f}'
                else:
                    s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 12), 48)

        # --- Freeze header + auto-filter ---
        ws.freeze_panes = 'A2'
        if n_data:
            ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}{1 + n_data}'

        # --- Ringkasan di bawah tabel (dengan jarak 1 baris) ---
        if ringkasan:
            start = 3 + n_data
            tcell = ws.cell(row=start, column=1, value='Ringkasan')
            tcell.font = Font(bold=True, size=12)
            for i, (label, nilai) in enumerate(ringkasan, start=start + 1):
                lc = ws.cell(row=i, column=1, value=label)
                lc.font = Font(bold=True)
                vc = ws.cell(row=i, column=2, value=nilai)
                if isinstance(nilai, (int, float)) and not isinstance(nilai, bool):
                    vc.number_format = MONEY_FMT
                    vc.alignment = Alignment(horizontal='right')

        wb.save(response)

    def _html(self, entry, columns, rows, ringkasan=None):
        """PDF via halaman HTML yang otomatis membuka dialog cetak — pola yang
        sama dipakai ExportCustomerNotesView, tanpa dependensi baru."""
        head = ''.join(f'<th>{c["label"]}</th>' for c in columns)
        body = ''.join(
            '<tr>' + ''.join(f'<td>{r.get(c["key"], "")}</td>' for c in columns) + '</tr>'
            for r in rows
        )
        if not rows:
            body = f'<tr><td colspan="{len(columns)}" style="text-align:center;color:#64748b">Tidak ada data.</td></tr>'

        blok_ringkasan = ''
        if ringkasan:
            baris = ''.join(
                f'<tr><th style="width:240px">{label}</th><td>{nilai}</td></tr>'
                for label, nilai in ringkasan
            )
            blok_ringkasan = f'<h3>Ringkasan</h3><table class="ringkasan"><tbody>{baris}</tbody></table>'

        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>{entry['label']}</title>
<style>body{{font-family:Arial,sans-serif;font-size:11px;padding:16px}}
h2{{font-size:15px;margin:0 0 10px}}
h3{{font-size:13px;margin:16px 0 6px}}
table{{width:100%;border-collapse:collapse}}
table.ringkasan{{width:auto;min-width:360px}}
th,td{{border:1px solid #cbd5e1;padding:5px 7px;text-align:left}}
th{{background:#f1f5f9}}</style></head><body>
<h2>{entry['label']}</h2><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>
{blok_ringkasan}
<script>window.onload=function(){{setTimeout(function(){{window.print();}},400);}};</script>
</body></html>"""
        return HttpResponse(html, content_type='text/html')
