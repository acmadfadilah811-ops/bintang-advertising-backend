"""View dashboard eksekutif.

Dipisah dari logika agregasi (executive_dashboard.py) mengikuti pola
production_views.py / production_costing.py.
"""

from django.http import HttpResponse
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from api.permissions import IsOwnerOrManager

from . import executive_dashboard

PERIODE_VALID = {'mtd', 'qtd', 'ytd', '12m'}


def _period(request):
    """Validasi periode. Menolak nilai asing daripada diam-diam memakai default."""
    period = request.query_params.get('period', 'ytd')
    if period not in PERIODE_VALID:
        return None, Response(
            {'error': f"Periode '{period}' tidak dikenal. Pilihan: {', '.join(sorted(PERIODE_VALID))}."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    return period, None


class ExecutiveDashboardView(APIView):
    """GET /api/executive-dashboard/?period=mtd|qtd|ytd|12m

    Ringkasan keuangan seluruh perusahaan — dibatasi owner/manager. Staff dan
    kasir tidak boleh melihat omzet, HPP, dan piutang agregat.
    """

    permission_classes = [IsAuthenticated, IsOwnerOrManager]

    def get(self, request):
        period, error = _period(request)
        if error:
            return error
        return Response(executive_dashboard.build(period))


class ExecutiveDashboardExportView(APIView):
    """GET /api/executive-dashboard/export/ — unduh XLSX."""

    permission_classes = [IsAuthenticated, IsOwnerOrManager]

    def get(self, request):
        period, error = _period(request)
        if error:
            return error

        # Diimpor di dalam method: openpyxl hanya dibutuhkan saat ekspor, jadi
        # kegagalan impor tidak ikut menjatuhkan endpoint overview.
        from openpyxl import Workbook

        data = executive_dashboard.build(period)
        wb = Workbook()

        ws = wb.active
        ws.title = 'Ringkasan'
        ws.append(['Periode', data['periode']['label']])
        ws.append([])
        ws.append(['KPI', 'Nilai', 'Perubahan (%)'])
        for item in data['kpi']:
            ws.append([item['label'], item['value'], item['delta'] if item['delta'] is not None else '-'])
        ws.append([])
        ws.append(['Belum tersedia', 'Alasan'])
        for item in data['unavailable']:
            ws.append([item['label'], item['reason']])

        ws2 = wb.create_sheet('Tren Bulanan')
        ws2.append(['Periode', 'Pendapatan', 'HPP', 'Laba Kotor'])
        for row in data['tren']:
            ws2.append([row['periode'], row['pendapatan'], row['hpp'], row['laba_kotor']])

        ws3 = wb.create_sheet('Produk Terlaris')
        ws3.append(['Produk', 'Qty', 'Nilai'])
        for row in data['produk_terlaris']:
            ws3.append([row['nama'], row['qty'], row['nilai']])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="dashboard-eksekutif-{data["periode"]["mulai"]}.xlsx"'
        )
        wb.save(response)
        return response
