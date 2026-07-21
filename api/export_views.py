import openpyxl
import logging
from django.http import HttpResponse, Http404
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from api.permissions import IsOwnerOrManager
from rest_framework.response import Response
from rest_framework.negotiation import DefaultContentNegotiation
from .models import Order, InventoryItem, JobBoard, Contact
from .customer_models import Customer
from django.utils import timezone

logger = logging.getLogger(__name__)


class IgnoreFormatContentNegotiation(DefaultContentNegotiation):
    def filter_renderers(self, renderers, format):
        try:
            return super().filter_renderers(renderers, format)
        except Http404:
            return renderers

class ExportContactsView(APIView):
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="pelanggan_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Pelanggan"

            headers = ['Nama', 'No. WhatsApp', 'Total Order', 'Total Belanja (Rp)', 'Terakhir Order', 'Produk Dipesan', 'Keterangan']
            ws.append(headers)

            contacts = Contact.objects.all().order_by('-total_spent')
            
            # Optimasi N+1: Prefetch seluruh order sekaligus dan kelompokkan menggunakan dictionary
            from collections import defaultdict
            orders_by_wa = defaultdict(list)
            all_orders = Order.objects.prefetch_related('items').all().order_by('-waktu')
            for order in all_orders:
                orders_by_wa[order.nomor_wa].append(order)

            for c in contacts:
                # Ambil daftar produk yang dipesan beserta tanggalnya
                ordered_products_with_dates = []
                customer_orders = orders_by_wa.get(c.nomor_wa, [])
                for order in customer_orders:
                    tgl_str = order.waktu.strftime('%d/%m/%Y')
                    for item in order.items.all():
                        ordered_products_with_dates.append(f"{item.jenis_produk} ({tgl_str})")
                products_str = ", ".join(ordered_products_with_dates) if ordered_products_with_dates else "-"

                last_order_str = c.last_order.strftime('%Y-%m-%d') if hasattr(c.last_order, 'strftime') else str(c.last_order) if c.last_order else ''
                ws.append([
                    c.nama,
                    c.nomor_wa,
                    c.total_order,
                    c.total_spent,
                    last_order_str,
                    products_str,
                    c.keterangan or '',
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportCustomersView(APIView):
    """GET /api/export/customers/ — ekspor data Customer (Pelanggan & Supplier > Pelanggan)."""
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="daftar_pelanggan_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Daftar Pelanggan"

            headers = [
                'Kode Pelanggan', 'Nama', 'Tipe Pelanggan', 'Handphone', 'Email',
                'Jenis Kelamin', 'Tanggal Lahir', 'Nama Perusahaan', 'Batas Kredit',
                'Deposit', 'Loyalty Points', 'Terima Buletin', 'Status',
                'Tanggal Berakhir', 'Alamat', 'Kota', 'Kecamatan', 'Kode Pos', 'Catatan',
            ]
            ws.append(headers)

            customers = Customer.objects.select_related('customer_group').order_by('-created_at')
            for c in customers:
                ws.append([
                    c.kode_pelanggan,
                    c.nama,
                    c.customer_group.nama if c.customer_group else '-',
                    c.handphone,
                    c.email,
                    c.get_jenis_kelamin_display() if c.jenis_kelamin else '-',
                    c.tanggal_lahir.strftime('%Y-%m-%d') if c.tanggal_lahir else '-',
                    c.nama_perusahaan or '-',
                    float(c.batas_kredit),
                    float(c.deposit),
                    c.loyalty_points,
                    'Ya' if c.terima_buletin else 'Tidak',
                    'Dibekukan' if c.bekukan else 'Aktif',
                    c.tanggal_berakhir.strftime('%Y-%m-%d') if c.tanggal_berakhir else '-',
                    c.alamat,
                    c.kota,
                    c.kecamatan,
                    c.kode_pos,
                    c.catatan,
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportOrdersView(APIView):
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        # Base Query
        orders_qs = Order.objects.prefetch_related(
            'items', 'items__jobs', 'items__jobs__pic_staff'
        ).order_by('-waktu')

        if start_date_str:
            try:
                from django.utils.dateparse import parse_date
                s_date = parse_date(start_date_str)
                if s_date:
                    orders_qs = orders_qs.filter(waktu__date__gte=s_date)
            except Exception as e:
                logger.warning(f"Gagal mem-parse start_date '{start_date_str}': {e}")
        if end_date_str:
            try:
                from django.utils.dateparse import parse_date
                e_date = parse_date(end_date_str)
                if e_date:
                    orders_qs = orders_qs.filter(waktu__date__lte=e_date)
            except Exception as e:
                logger.warning(f"Gagal mem-parse end_date '{end_date_str}': {e}")

        # Safety limit to prevent memory issue
        total_count = orders_qs.count()
        if total_count > 50000:
            return Response({
                'error': 'Terlalu banyak data untuk diekspor. Gunakan filter rentang tanggal (start_date & end_date) untuk membatasi hasil.',
                'suggestion': f'Total data saat ini: {total_count} (Maksimal: 50.000)'
            }, status=400)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="orders_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Orders"

            headers = [
                'ID Pesanan', 'Tanggal', 'Nama Pelanggan', 'No WA', 
                'Produk yang Dipesan', 'Jumlah', 'PIC / Order Service', 
                'Status', 'Total Harga', 'Catatan'
            ]
            ws.append(headers)

            for order in orders_qs:
                items = order.items.all()
                products = ", ".join([it.jenis_produk for it in items])
                quantities = ", ".join([str(it.qty) for it in items])
                
                pics = set()
                for it in items:
                    for job in it.jobs.all():
                        if job.pic_staff:
                            pics.add(job.pic_staff.username)
                pic_str = ", ".join(pics) if pics else "-"

                total_harga = sum(item.harga_jual for item in items)
                ws.append([
                    order.id,
                    order.waktu.strftime('%Y-%m-%d %H:%M'),
                    order.nama,
                    order.nomor_wa,
                    products,
                    quantities,
                    pic_str,
                    order.get_status_global_display(),
                    total_harga,
                    order.catatan_pelanggan or ''
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportInventoryView(APIView):
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="inventory_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory"

            headers = ['ID Item', 'Nama', 'Kategori', 'Stok Saat Ini', 'Satuan', 'Min Stok', 'Harga Beli/Unit', 'Supplier', 'Tanggal Restok Terakhir', 'Tanggal Update/Mutasi', 'Status']
            ws.append(headers)

            items = InventoryItem.objects.prefetch_related('history').order_by('nama')
            for item in items:
                status = 'Kritis' if item.stok < item.min_stok else 'Aman'
                
                # Evaluasi di memory Python agar tidak memicu N+1 query.
                # Karena RestockHistory memiliki Meta ordering = ['-waktu'], list ini sudah berurutan terbalik secara default.
                history_list = list(item.history.all())
                
                latest_history = history_list[0] if history_list else None
                tanggal_update_str = latest_history.waktu.strftime('%Y-%m-%d %H:%M') if latest_history else '-'
                
                restock_list = [h for h in history_list if h.delta > 0]
                latest_restock = restock_list[0] if restock_list else None
                tanggal_restok_str = latest_restock.waktu.strftime('%Y-%m-%d %H:%M') if latest_restock else '-'

                ws.append([
                    item.id,
                    item.nama,
                    item.kategori,
                    item.stok,
                    item.satuan,
                    item.min_stok,
                    item.cost_per_unit,
                    item.supplier,
                    tanggal_restok_str,
                    tanggal_update_str,
                    status
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportJobsView(APIView):
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="jobs_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Jobs"

            headers = ['ID Job', 'ID Order', 'Jenis Produk', 'Tahap', 'Divisi', 'Staff PIC', 'Status Pekerjaan', 'Waktu Mulai', 'Waktu Selesai', 'Insentif']
            ws.append(headers)

            # FIX: select_related untuk semua relasi yang diakses dalam loop → eliminasi N+1 query berganda
            jobs = JobBoard.objects.select_related(
                'order_item__order',
                'tahap__divisi',
                'pic_staff'
            ).order_by('-id')

            for job in jobs:
                ws.append([
                    job.id,
                    job.order_item.order.id if job.order_item and job.order_item.order else '',
                    job.order_item.jenis_produk if job.order_item else '',
                    job.tahap.nama if job.tahap else 'Tanpa Tahap',
                    job.tahap.divisi.nama if job.tahap and job.tahap.divisi else 'Tanpa Divisi',
                    job.pic_staff.username if job.pic_staff else 'Belum diset',
                    job.get_status_pekerjaan_display(),
                    job.waktu_mulai.strftime('%Y-%m-%d %H:%M') if job.waktu_mulai else '-',
                    job.waktu_selesai.strftime('%Y-%m-%d %H:%M') if job.waktu_selesai else '-',
                    job.insentif
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response

class ExportAbsensiView(APIView):
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        tanggal = request.query_params.get("tanggal")
        if not tanggal:
            tanggal = timezone.localdate().strftime("%Y-%m-%d")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="absensi_{tanggal}.xlsx"'

        try:
            from hr.models import Absensi
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Absensi {tanggal}"

            headers = ['Nama Staff', 'Divisi', 'Status Kehadiran', 'Waktu Masuk', 'Waktu Keluar', 'Durasi Kerja (Jam)', 'Keterangan']
            ws.append(headers)

            absensi = Absensi.objects.filter(tanggal=tanggal).select_related('staff', 'staff__divisi').order_by('staff__username')

            for a in absensi:
                masuk = a.jam_masuk.strftime('%H:%M:%S') if a.jam_masuk else '-'
                keluar = a.jam_keluar.strftime('%H:%M:%S') if a.jam_keluar else '-'
                ws.append([
                    a.staff.get_full_name() or a.staff.username,
                    a.staff.divisi.nama if a.staff.divisi else 'Belum Ditentukan',
                    a.get_status_display(),
                    masuk,
                    keluar,
                    a.durasi_kerja_jam,
                    a.catatan or ''
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportStaffPerformanceView(APIView):
    """
    GET /api/export/staff-performance/?range=bulan_ini&type=global&divisi=...&staff_id=...
    Mengekspor laporan detail kinerja (global/divisi/staff) ke Excel.
    """
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        time_range = request.query_params.get('range', 'bulan_ini')
        export_type = request.query_params.get('type', 'global')  # 'global', 'divisi', 'staff'
        divisi_name = request.query_params.get('divisi', '')
        staff_id = request.query_params.get('staff_id', '')

        import calendar
        from api.models import JobBoard, CustomUser
        
        now = timezone.localtime(timezone.now())
        start = None
        end = None

        if time_range == 'hari_ini':
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif time_range == 'tujuh_hari':
            start = (now - timezone.timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
            end = now.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif time_range == 'bulan_ini':
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            _, last_day = calendar.monthrange(now.year, now.month)
            end = now.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)
        elif time_range == 'bulan_lalu':
            year = now.year
            month = now.month - 1
            if month <= 0:
                month = 12
                year -= 1
            start = now.replace(year=year, month=month, day=1, hour=0, minute=0, second=0, microsecond=0)
            _, last_day = calendar.monthrange(year, month)
            end = now.replace(year=year, month=month, day=last_day, hour=23, minute=59, second=59, microsecond=999999)

        jobs_qs = JobBoard.objects.select_related(
            'order_item', 'order_item__order', 'tahap', 'tahap__divisi', 'pic_staff'
        ).order_by('-waktu_selesai')

        if start and end:
            jobs_qs = jobs_qs.filter(waktu_selesai__gte=start, waktu_selesai__lte=end)
        else:
            jobs_qs = jobs_qs.filter(waktu_selesai__isnull=False)

        title_suffix = "Global"
        filename_prefix = "laporan_global"

        if export_type == 'divisi' and divisi_name:
            jobs_qs = jobs_qs.filter(tahap__divisi__nama__iexact=divisi_name)
            title_suffix = f"Divisi {divisi_name}"
            filename_prefix = f"laporan_divisi_{divisi_name.lower().replace(' ', '_')}"
        elif export_type == 'staff' and staff_id:
            try:
                staff_user = CustomUser.objects.get(id=staff_id)
                jobs_qs = jobs_qs.filter(pic_staff=staff_user)
                title_suffix = f"Staff {staff_user.get_full_name() or staff_user.username}"
                filename_prefix = f"laporan_staff_{staff_user.username}"
            except CustomUser.DoesNotExist:
                return Response({'error': 'Staff tidak ditemukan'}, status=404)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{time_range}_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Detail Pekerjaan"

            # Title block
            ws.append([f"LAPORAN DETAIL PRODUKSI - {title_suffix.upper()}"])
            periode_str = f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}" if start else "Semua Waktu"
            ws.append([f"Periode: {time_range.replace('_', ' ').title()} ({periode_str})"])
            ws.append([]) # Blank spacer

            headers = [
                'No', 'ID Job', 'Tanggal Selesai', 'ID Order', 'Pelanggan', 
                'Nama Produk', 'Bahan Baku', 'Qty', 'Divisi', 'Tahap Proses', 
                'ID PIC', 'Nama PIC', 'Durasi (Menit)', 'Status', 'Insentif (Rp)', 'Biaya Desain (Rp)'
            ]
            ws.append(headers)

            for idx, job in enumerate(jobs_qs, 1):
                duration_minutes = 0
                if job.waktu_mulai and job.waktu_selesai:
                    diff = job.waktu_selesai - job.waktu_mulai
                    duration_minutes = round(diff.total_seconds() / 60.0, 1)

                ws.append([
                    idx,
                    job.id,
                    job.waktu_selesai.strftime('%Y-%m-%d %H:%M') if job.waktu_selesai else '-',
                    job.order_item.order.id if job.order_item and job.order_item.order else '-',
                    job.order_item.order.nama if job.order_item and job.order_item.order else '-',
                    job.order_item.jenis_produk if job.order_item else '-',
                    job.order_item.bahan if job.order_item else '-',
                    job.order_item.qty if job.order_item else 1,
                    job.tahap.divisi.nama if job.tahap and job.tahap.divisi else '-',
                    job.tahap.nama if job.tahap else '-',
                    job.pic_staff.username if job.pic_staff else '-',
                    job.pic_staff.get_full_name() or job.pic_staff.username if job.pic_staff else '-',
                    duration_minutes,
                    job.status_pekerjaan,
                    job.insentif,
                    job.biaya_desain
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportStockMovementView(APIView):
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        import datetime
        from django.utils.dateparse import parse_date

        if start_date_str:
            start_date = parse_date(start_date_str)
        else:
            start_date = datetime.date.today()

        if end_date_str:
            end_date = parse_date(end_date_str)
        else:
            end_date = datetime.date.today()

        filename = f"summary-{start_date.strftime('%Y-%m-%d')}__{end_date.strftime('%Y-%m-%d')}.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pergerakan Stok"

            headers = ['Grup', 'Produk', 'Awal', 'Masuk', 'Pengembalian', 'Penjualan', 'Keluar', 'Sisa']
            ws.append(headers)

            from .product_models import Product, ProductVariant, ProductStockMovement

            products = Product.objects.all().select_related('kategori').prefetch_related('variants')

            skus = {}
            for p in products:
                if p.has_variant and p.variants.exists():
                    for v in p.variants.all():
                        skus[(p.id, v.id)] = {
                            'product_id': p.id,
                            'variant_id': v.id,
                            'product_name': p.nama,
                            'variant_name': v.nama_varian,
                            'sku': v.sku or p.sku or '',
                            'group': p.kategori.nama if p.kategori else 'Umum',
                            'current_qty': v.qty_stok,
                        }
                else:
                    skus[(p.id, None)] = {
                        'product_id': p.id,
                        'variant_id': None,
                        'product_name': p.nama,
                        'variant_name': '',
                        'sku': p.sku or '',
                        'group': p.kategori.nama if p.kategori else 'Umum',
                        'current_qty': p.qty_stok,
                    }

            movements = ProductStockMovement.objects.all().order_by('created_at')

            summary_map = {}
            for key, info in skus.items():
                summary_map[key] = {
                    'group': info['group'],
                    'product': f"{info['product_name']} - {info['variant_name']}" if info['variant_name'] else info['product_name'],
                    'initial': float(info['current_qty']),
                    'in': 0.0,
                    'returnStock': 0.0,
                    'sales': 0.0,
                    'out': 0.0,
                    'sisa': float(info['current_qty']),
                    'movements_before': [],
                    'movements_during': [],
                }

            for m in movements:
                m_date = m.tanggal or m.created_at.date()
                key = (m.product_id, m.variant_id)
                if key not in summary_map:
                    continue

                if m_date < start_date:
                    summary_map[key]['movements_before'].append(m)
                elif start_date <= m_date <= end_date:
                    summary_map[key]['movements_during'].append(m)

            for key, s in summary_map.items():
                info = skus[key]

                if s['movements_before']:
                    last_before = s['movements_before'][-1]
                    initial_val = last_before.stok_akhir
                elif s['movements_during']:
                    first_during = s['movements_during'][0]
                    initial_val = first_during.stok_awal
                else:
                    movements_after = ProductStockMovement.objects.filter(
                        product_id=info['product_id'],
                        variant_id=info['variant_id']
                    ).order_by('created_at')
                    after_list = []
                    for m in movements_after:
                        m_date = m.tanggal or m.created_at.date()
                        if m_date > end_date:
                            after_list.append(m)
                    if after_list:
                        initial_val = after_list[0].stok_awal
                    else:
                        initial_val = info['current_qty']

                s['initial'] = float(initial_val)

                in_qty = 0.0
                out_qty = 0.0
                sales_qty = 0.0
                return_qty = 0.0

                for m in s['movements_during']:
                    qty = float(m.qty)
                    if m.tipe in ('masuk', 'produksi'):
                        in_qty += qty
                    elif m.tipe == 'keluar':
                        out_qty += qty
                    elif m.tipe == 'penjualan':
                        sales_qty += qty
                    elif m.tipe == 'pengembalian':
                        return_qty += qty

                s['in'] = in_qty
                s['out'] = out_qty
                s['sales'] = sales_qty
                s['returnStock'] = return_qty

                if s['movements_during']:
                    sisa_val = s['movements_during'][-1].stok_akhir
                elif s['movements_before']:
                    sisa_val = s['movements_before'][-1].stok_akhir
                else:
                    sisa_val = initial_val

                s['sisa'] = float(sisa_val)

                ws.append([
                    s['group'],
                    s['product'],
                    s['initial'],
                    s['in'],
                    s['returnStock'],
                    s['sales'],
                    s['out'],
                    s['sisa'],
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportProductsView(APIView):
    permission_classes = [IsOwnerOrManager]

    def get(self, request):
        import csv
        from django.http import HttpResponse
        from .product_models import Product

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="products_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        # Ensure correct UTF-8 BOM encoding for Excel compatibility
        response.write(u'\ufeff'.encode('utf8'))
        
        writer = csv.writer(response)
        
        # Headers
        headers = [
            "name", "alternative_name", "classification_id", "category",
            "variant_column_labels", "variant_names", "alternative_variant_name",
            "variant_view_order", "collections", "brand", "condition_id", "uom",
            "sku", "barcode", "serial_nos", "description", "buy_price",
            "sell_price", "pos_sell_price", "market_price", "wholesale_group",
            "wholesale_price", "pos_sell_price_dynamic", "track_inventory",
            "stock_qty", "low_stock_alert", "is_out_stock", "weight_kg",
            "loyalty_points", "comission", "published", "hidden_in_pos",
            "qty_fast_moving", "rack", "is_always_available", "non_taxable",
            "customer_comission", "is_customer_comission_percentage", "non_service_charge"
        ]
        writer.writerow(headers)

        products = Product.objects.select_related('kategori', 'brand', 'koleksi').prefetch_related('variants').all()
        
        for p in products:
            category_name = p.kategori.nama if p.kategori else ""
            brand_name = p.brand.nama if p.brand else ""
            collection_name = p.koleksi.nama if p.koleksi else ""
            
            if p.has_variant and p.variants.exists():
                for idx, v in enumerate(p.variants.all()):
                    writer.writerow([
                        p.nama,
                        p.nama_alternatif or "",
                        "",  # classification_id
                        category_name,
                        "Varian",  # variant_column_labels
                        v.nama_varian,
                        v.nama_alternatif or "",
                        str(idx + 1),  # variant_view_order
                        collection_name,
                        brand_name,
                        "N",  # condition_id
                        p.satuan or "pcs",
                        v.sku or "",
                        v.barcode or "",
                        "",  # serial_nos
                        p.deskripsi or "",
                        str(v.harga_beli),
                        str(v.harga_jual_online),
                        str(v.harga_jual_toko),
                        str(v.harga_pasar),
                        "",  # wholesale_group
                        "0",  # wholesale_price
                        "0",  # pos_sell_price_dynamic
                        "1" if v.lacak_inventori else "0",
                        str(v.qty_stok),
                        str(p.stok_minimum),
                        "0" if v.qty_stok > 0 else "1",
                        str(float(v.berat or 0.0) / 1000.0),  # weight_kg
                        "0",  # loyalty_points
                        "0",  # commission
                        "1" if p.is_active else "0",
                        "0",  # hidden_in_pos
                        str(p.qty_fast_moving),
                        v.rack or p.rack or "",
                        "1" if not v.lacak_inventori else "0",
                        "0",  # non_taxable
                        "0",  # customer_comission
                        "0",  # is_customer_comission_percentage
                        "0"   # non_service_charge
                    ])
            else:
                writer.writerow([
                    p.nama,
                    p.nama_alternatif or "",
                    "",  # classification_id
                    category_name,
                    "",  # variant_column_labels
                    "",  # variant_names
                    "",  # alternative_variant_name
                    "",  # variant_view_order
                    collection_name,
                    brand_name,
                    "N",  # condition_id
                    p.satuan or "pcs",
                    p.sku or "",
                    p.barcode or "",
                    "",  # serial_nos
                    p.deskripsi or "",
                    str(p.harga_beli),
                    str(p.harga_jual_online),
                    str(p.harga_jual_toko),
                    "0",  # market_price
                    "",  # wholesale_group
                    "0",  # wholesale_price
                    "0",  # pos_sell_price_dynamic
                    "1" if p.lacak_inventori else "0",
                    str(p.qty_stok),
                    str(p.stok_minimum),
                    "0" if p.qty_stok > 0 else "1",
                    "0",  # weight_kg
                    "0",  # loyalty_points
                    "0",  # commission
                    "1" if p.is_active else "0",
                    "0",  # hidden_in_pos
                    str(p.qty_fast_moving),
                    p.rack or "",
                    "1" if not p.lacak_inventori else "0",
                    "0",  # non_taxable
                    "0",  # customer_comission
                    "0",  # is_customer_comission_percentage
                    "0"   # non_service_charge
                ])
                
        return response


class ExportCustomerNotesView(APIView):
    permission_classes = [IsOwnerOrManager]
    content_negotiation_class = IgnoreFormatContentNegotiation

    def get(self, request):
        from .customer_models import CustomerNote
        
        # Get query parameters (filters)
        customer_id = request.query_params.get('customer')
        tag_name = request.query_params.get('tag')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        search_query = request.query_params.get('search')
        export_format = request.query_params.get('format', 'excel')

        # Base query
        notes_qs = CustomerNote.objects.select_related('customer', 'dibuat_oleh').prefetch_related('tags', 'entries').order_by('-created_at')

        # Apply filters
        if customer_id:
            notes_qs = notes_qs.filter(customer_id=customer_id)
        if tag_name:
            notes_qs = notes_qs.filter(tags__nama=tag_name)
        if start_date:
            notes_qs = notes_qs.filter(tanggal__gte=start_date)
        if end_date:
            notes_qs = notes_qs.filter(tanggal__lte=end_date)
        if search_query:
            from django.db.models import Q
            notes_qs = notes_qs.filter(
                Q(judul__icontains=search_query) |
                Q(customer_name__icontains=search_query)
            )

        # Safety limit (consistent with system architecture)
        total_count = notes_qs.count()
        if total_count > 50000:
            return Response({
                'error': f'Data terlalu banyak ({total_count:,} records). Limit: 50,000',
            }, status=400)

        if export_format == 'pdf':
            # Generate a beautiful print-ready HTML page
            html_content = self.generate_pdf_html(notes_qs)
            response = HttpResponse(html_content, content_type='text/html')
            return response

        # Default format: Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="customer-notes-{timezone.now().strftime("%Y-%m-%d")}.xlsx"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Catatan Pelanggan"

            headers = ['id', 'customer_id', 'customer_name', 'note_no', 'title', 'note', 'created_time', 'tag', 'created_by_name']
            ws.append(headers)

            for note in notes_qs:
                note_no = f"N{note.id:06d}"
                created_time = f"{note.tanggal.strftime('%Y-%m-%d') if note.tanggal else ''} {note.jam.strftime('%H:%M') if note.jam else ''}".strip()
                if not created_time and note.created_at:
                    created_time = timezone.localtime(note.created_at).strftime('%Y-%m-%d %H:%M')

                # note column: concatenate all entry contents
                entries_list = []
                for entry in note.entries.all():
                    val = f"[{entry.label}] {entry.content}" if entry.label else entry.content
                    entries_list.append(val)
                note_text = "\n".join(entries_list) if entries_list else None

                # tag column: list tags separated by comma
                tags_list = [t.nama for t in note.tags.all()]
                tag_text = ", ".join(tags_list) if tags_list else None

                created_by = note.dibuat_oleh.get_full_name() or note.dibuat_oleh.username if note.dibuat_oleh else 'Brandy'

                ws.append([
                    note.id,
                    note.customer_id or '',
                    note.customer_name or 'Umum / Anonim',
                    note_no,
                    note.judul or '',
                    note_text,
                    created_time,
                    tag_text,
                    created_by
                ])
            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response

    def generate_pdf_html(self, notes_qs):
        import datetime
        now_str = datetime.datetime.now().strftime('%d %B %Y, %H:%M')
        
        # Build rows
        rows_html = ""
        for idx, note in enumerate(notes_qs, 1):
            note_no = f"N{note.id:06d}"
            created_time = f"{note.tanggal.strftime('%Y-%m-%d') if note.tanggal else ''} {note.jam.strftime('%H:%M') if note.jam else ''}".strip()
            if not created_time and note.created_at:
                created_time = timezone.localtime(note.created_at).strftime('%Y-%m-%d %H:%M')

            # entries
            entries_list = []
            for entry in note.entries.all():
                val = f"<strong>[{entry.label}]</strong> {entry.content}" if entry.label else entry.content
                entries_list.append(val)
            note_text = "<br>".join(entries_list) if entries_list else "-"

            # tags
            tags_list = [f'<span class="badge">{t.nama}</span>' for t in note.tags.all()]
            tags_html = " ".join(tags_list) if tags_list else "-"

            created_by = note.dibuat_oleh.get_full_name() or note.dibuat_oleh.username if note.dibuat_oleh else 'Brandy'

            rows_html += f"""
            <tr>
                <td style="text-align: center;">{idx}</td>
                <td><strong>{note.judul or '(Tanpa judul)'}</strong><br><small style="color: #64748b;">{note_no}</small></td>
                <td>{note.customer_name or 'Umum / Anonim'}</td>
                <td style="white-space: nowrap;">{created_time}</td>
                <td>{tags_html}</td>
                <td>{created_by}</td>
                <td style="font-size: 12px; line-height: 1.4;">{note_text}</td>
            </tr>
            """

        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Laporan Catatan Pelanggan</title>
    <style>
        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            color: #1e293b;
            margin: 20px;
            font-size: 13px;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 2px solid #e2e8f0;
            padding-bottom: 15px;
            margin-bottom: 20px;
        }}
        .title {{
            font-size: 20px;
            font-weight: bold;
            color: #0f172a;
        }}
        .meta-info {{
            font-size: 12px;
            color: #64748b;
            margin-bottom: 20px;
            line-height: 1.5;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }}
        th, td {{
            border: 1px solid #cbd5e1;
            padding: 8px 10px;
            text-align: left;
            vertical-align: top;
        }}
        th {{
            background-color: #f1f5f9;
            color: #334155;
            font-weight: bold;
        }}
        .badge {{
            display: inline-block;
            padding: 2px 6px;
            border-radius: 4px;
            background-color: #eff6ff;
            color: #1d4ed8;
            font-size: 10px;
            font-weight: bold;
            border: 1px solid #dbeafe;
            margin-right: 2px;
            margin-bottom: 2px;
        }}
        .no-print-btn {{
            background-color: #2563eb;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 6px;
            font-size: 13px;
            font-weight: bold;
            cursor: pointer;
            margin-bottom: 20px;
        }}
        .no-print-btn:hover {{
            background-color: #1d4ed8;
        }}
        @media print {{
            .no-print {{
                display: none !important;
            }}
            body {{
                margin: 0;
                color: #000;
            }}
            th {{
                background-color: #f8fafc !important;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
        }}
    </style>
</head>
<body>
    <div class="no-print" style="display: flex; justify-content: space-between; align-items: center;">
        <button class="no-print-btn" onclick="window.print()">Cetak Halaman (PDF)</button>
        <span style="color: #64748b; font-size: 12px;">Tips: Atur printer ke "Save as PDF" untuk mengunduh dokumen.</span>
    </div>

    <div class="header">
        <div>
            <div class="title">LAPORAN CATATAN PELANGGAN</div>
            <div style="font-size: 13px; color: #475569; margin-top: 4px;">Bintang Advertising ERP</div>
        </div>
        <div style="text-align: right; font-size: 12px; color: #64748b;">
            Waktu Cetak: {now_str}
        </div>
    </div>

    <div class="meta-info">
        <strong>Status Data:</strong> Diekspor dari database sistem<br>
        <strong>Total Catatan:</strong> {len(notes_qs)} data ditemukan
    </div>

    <table>
        <thead>
            <tr>
                <th style="width: 30px; text-align: center;">No</th>
                <th style="width: 120px;">Judul & ID</th>
                <th style="width: 120px;">Pembeli</th>
                <th style="width: 110px;">Waktu Buat</th>
                <th style="width: 120px;">Tag</th>
                <th style="width: 100px;">Dibuat Oleh</th>
                <th>Catatan / Entri</th>
            </tr>
        </thead>
        <tbody>
            {rows_html if rows_html else '<tr><td colspan="7" style="text-align: center; color: #64748b;">Tidak ada data catatan.</td></tr>'}
        </tbody>
    </table>

    <script>
        window.onload = function() {{
            setTimeout(function() {{
                window.print();
            }}, 500);
        }};
    </script>
</body>
</html>
"""
        return html

class ExportCashTransactionsView(APIView):
    """Export Excel untuk Pendapatan/Pengeluaran (Kas Masuk/Keluar).

    Menghormati filter yang sama dengan layar: ?arah=, ?start=, ?end=, ?search=
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .finance_models import CashTransaction

        qs = CashTransaction.objects.all().select_related('tipe_transaksi', 'staff')

        arah = request.query_params.get('arah')
        if arah in ('pendapatan', 'pengeluaran'):
            qs = qs.filter(arah=arah)

        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if start:
            qs = qs.filter(waktu__date__gte=start)
        if end:
            qs = qs.filter(waktu__date__lte=end)

        search = (request.query_params.get('search') or '').strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(nomor__icontains=search)
                | Q(catatan__icontains=search)
                | Q(tipe_transaksi__nama__icontains=search)
                | Q(staff__username__icontains=search)
            )

        qs = qs.order_by('-waktu')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="pendapatan_pengeluaran_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pendapatan-Pengeluaran"

            ws.append(['No. Transaksi', 'Tanggal', 'Waktu', 'Staff', 'Tipe Transaksi', 'Arah', 'Jumlah (Rp)', 'Catatan'])

            total_masuk = 0
            total_keluar = 0
            for t in qs:
                waktu_local = timezone.localtime(t.waktu) if timezone.is_aware(t.waktu) else t.waktu
                jumlah = float(t.jumlah or 0)
                if t.arah == 'pendapatan':
                    total_masuk += jumlah
                else:
                    total_keluar += jumlah
                ws.append([
                    t.nomor,
                    waktu_local.strftime('%Y-%m-%d'),
                    waktu_local.strftime('%H:%M'),
                    t.staff.username if t.staff else '',
                    t.tipe_transaksi.nama if t.tipe_transaksi else '',
                    'Pendapatan' if t.arah == 'pendapatan' else 'Pengeluaran',
                    jumlah,
                    t.catatan or '',
                ])

            # Ringkasan di bawah tabel
            ws.append([])
            ws.append(['', '', '', '', '', 'Total Pendapatan', total_masuk, ''])
            ws.append(['', '', '', '', '', 'Total Pengeluaran', total_keluar, ''])
            ws.append(['', '', '', '', '', 'Selisih (Net)', total_masuk - total_keluar, ''])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportSalesItemsByBrandView(APIView):
    """
    GET /api/export/sales-items-by-brand/?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD
    Mengekspor laporan Item Penjualan Berdasarkan Brand (27 kolom legacy format).
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from api.pos_models import POSSaleItem
        from api.models import OrderItem
        from django.utils.dateparse import parse_date
        import datetime

        start_date_str = request.query_params.get('start_date') or request.query_params.get('start')
        end_date_str = request.query_params.get('end_date') or request.query_params.get('end')

        if start_date_str:
            start_date = parse_date(start_date_str)
        else:
            start_date = datetime.date.today()

        if end_date_str:
            end_date = parse_date(end_date_str)
        else:
            end_date = datetime.date.today()

        filename = f"Item Penjualan Berdasarkan Brand-{start_date.strftime('%Y-%m-%d')}__{end_date.strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Items by Brand"

            headers = [
                'Grup', 'Produk', 'Label Varian', 'Nama Alternatif', 'Merek', 'Diskon', 
                'Total Diskon', 'Penjualan Kotor', 'Diskon Produk', 'Penjualan Bersih', 
                'Harga Beli', 'Total Harga Beli', 'Laba Kotor', 'SKU', 'Barcode', 
                'Serial/IMEI', 'Jumlah Penjualan', 'Metode Penjualan', 'Cabang', 
                'Komisi Karyawan', 'Point Terpakai', 'Penerimaan Point', 'Penerimaan Komisi Pelanggan', 
                'Total Pengembalian', 'Pengembalian', 'Diskon Pengembalian', 'Pengembalian Kotor'
            ]
            ws.append(headers)

            # 1. Fetch POS Sale Items
            pos_items = POSSaleItem.objects.filter(
                sale__created_at__date__gte=start_date,
                sale__created_at__date__lte=end_date
            ).select_related('sale', 'product', 'product__brand', 'product__kategori', 'variant')

            for item in pos_items:
                product_name = item.nama_snapshot or (item.product.nama if item.product else '')
                variant_label = item.variant.nama_varian if item.variant else ''
                brand_name = item.product.brand.nama if item.product and item.product.brand else ''
                group_name = item.product.kategori.nama if item.product and item.product.kategori else 'Umum'
                sku = item.variant.sku if item.variant else (item.product.sku if item.product else '')
                barcode = item.variant.barcode if item.variant else (item.product.barcode if item.product else '')

                qty = float(item.qty or 1)
                unit_price = float(item.harga_snapshot or 0)
                subtotal = float(item.subtotal or (qty * unit_price))
                cost_price = float(item.product.harga_beli if item.product else 0)
                total_cost = cost_price * qty
                gross_profit = subtotal - total_cost

                ws.append([
                    group_name, product_name, variant_label, '', brand_name, 0,
                    0, subtotal, 0, subtotal,
                    cost_price, total_cost, gross_profit, sku, barcode,
                    '', qty, 'POS / Toko Direct', 'Bintang Advertising',
                    0, 0, 0, 0,
                    0, 0, 0, 0
                ])

            # 2. Fetch Production Order Items
            order_items = OrderItem.objects.filter(
                order__waktu__date__gte=start_date,
                order__waktu__date__lte=end_date
            ).select_related('order')

            for item in order_items:
                product_name = item.jenis_produk or 'Custom Printing'
                variant_label = f"{item.panjang}x{item.lebar}m ({item.bahan})" if item.bahan else ''
                qty = float(item.qty or 1)
                unit_price = float(item.harga_jual or 0)
                subtotal = qty * unit_price
                cost_price = float(item.harga_modal or 0)
                total_cost = cost_price * qty
                gross_profit = subtotal - total_cost

                ws.append([
                    'Produksi', product_name, variant_label, '', 'Bintang', 0,
                    0, subtotal, 0, subtotal,
                    cost_price, total_cost, gross_profit, '', '',
                    '', qty, 'Order Produksi', 'Bintang Advertising',
                    0, 0, 0, 0,
                    0, 0, 0, 0
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response


class ExportSalesDetailsView(APIView):
    """
    GET /api/export/sales-details/?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD
    Mengekspor laporan Rincian Penjualan (48 kolom legacy format).
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from api.pos_models import POSSale
        from api.models import Order
        from django.utils.dateparse import parse_date
        import datetime

        start_date_str = request.query_params.get('start_date') or request.query_params.get('start')
        end_date_str = request.query_params.get('end_date') or request.query_params.get('end')

        if start_date_str:
            start_date = parse_date(start_date_str)
        else:
            start_date = datetime.date.today()

        if end_date_str:
            end_date = parse_date(end_date_str)
        else:
            end_date = datetime.date.today()

        filename = f"Rincian Penjualan-{start_date.strftime('%Y-%m-%d')}__{end_date.strftime('%Y-%m-%d')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Details"

            headers = [
                'Waktu', 'No. Penjualan', 'Metode Penjualan', 'Metode Pembayaran', 
                'Total Pembayaran', 'Penerimaan Pembayaran', 'Kembalian Pembayaran', 
                'Penjualan Kotor', 'Penjualan Bersih', 'Diskon Produk', 'Diskon Penjualan', 
                'Voucher', 'Biaya Pengiriman', 'Service Charge', 'Laba Kotor', 
                'Kasir', 'Pelanggan', 'Grup Pelanggan', 'Biaya Pembayaran', 'Status', 
                'Catatan Pembatalan', 'Alasan Pembatalan', 'Tanggal Pengiriman', 'Status Pengiriman', 
                'Nomor Resi', 'Kurir', 'Kota', 'Provinsi', 'Kode Pos', 'Alamat', 
                'Nama Penerima', 'No. Telepon Penerima', 'Catatan Penjualan', 'Komisi Karyawan', 
                'Merek', 'Cabang', 'Penerimaan Point', 'Point Terpakai', 'Penerimaan Komisi Pelanggan', 
                'Pengembalian Bersih', 'Penjualan Setelah Pengembalian', 'Pengembalian Kotor', 
                'Pengembalian Diskon Produk', 'Pengembalian Diskon Penjualan', 'PPN', 'Tarif PPN (%)', 
                'Penjualan Sebelum PPN', 'Pengembalian PPN'
            ]
            ws.append(headers)

            # 1. Fetch POS Sales
            pos_sales = POSSale.objects.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).select_related('kasir', 'pelanggan')

            for s in pos_sales:
                waktu_str = s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else ''
                total_harga = float(s.total or 0)
                bayar = float(s.dibayar or 0)
                kembali = float(s.kembalian or 0)
                kasir_name = s.kasir.get_full_name() or s.kasir.username if s.kasir else 'Kasir POS'
                customer_name = s.pelanggan.nama if s.pelanggan else 'Pelanggan Umum'

                ws.append([
                    waktu_str, s.nomor, 'POS Direct', s.metode_bayar or 'Cash',
                    total_harga, bayar, kembali,
                    total_harga, total_harga, 0, float(s.diskon or 0),
                    0, 0, 0, total_harga * 0.3,
                    kasir_name, customer_name, 'Umum', 0, s.status,
                    '', '', '', '',
                    '', '', '', '', '', '',
                    customer_name, '', s.catatan or '', 0,
                    'Bintang', 'Bintang Advertising', 0, 0, 0,
                    0, total_harga, 0,
                    0, 0, float(s.pajak or 0), 0,
                    total_harga, 0
                ])

            # 2. Fetch Production Orders
            orders = Order.objects.filter(
                waktu__date__gte=start_date,
                waktu__date__lte=end_date
            )

            for o in orders:
                waktu_str = o.waktu.strftime('%Y-%m-%d %H:%M:%S') if o.waktu else ''
                total_harga = float(o.total_harga or 0)
                dp = float(o.dp_dibayar or 0)
                sisa = float(o.sisa_tagihan or 0)
                kasir_name = o.kasir.get_full_name() or o.kasir.username if o.kasir else 'Admin CS'
                customer_name = o.nama or 'Pelanggan Order'
                status_str = 'LUNAS' if sisa <= 0 else f'DP (Sisa {sisa:,.0f})'

                ws.append([
                    waktu_str, str(o.id), 'Order Produksi', o.metode_pembayaran or 'Transfer/Cash',
                    total_harga, dp, 0,
                    total_harga, total_harga, 0, float(o.diskon or 0),
                    0, 0, 0, total_harga * 0.35,
                    kasir_name, customer_name, 'Pelanggan Produksi', 0, status_str,
                    '', '', '', '',
                    '', '', '', '', '', '',
                    customer_name, o.nomor_wa or '', o.catatan_pelanggan or '', 0,
                    'Bintang', 'Bintang Advertising', 0, 0, 0,
                    0, total_harga, 0,
                    0, 0, 0, 0,
                    total_harga, 0
                ])

            wb.save(response)
        except Exception as e:
            return Response({'error': f'Gagal membuat file Excel: {str(e)}'}, status=500)

        return response

